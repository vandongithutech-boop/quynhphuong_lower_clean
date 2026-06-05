from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from transport.models import TransportRoute, TransportCompany, RouteCompany


def clean_value(value):
    if value is None:
        return ""

    value = str(value).strip()

    if value.lower() in ["nan", "none", "0"]:
        return ""

    return value


def detect_receive_type(factory_value, station_value):
    factory_value = clean_value(factory_value).upper()
    station_value = clean_value(station_value).upper()

    if "XƯỞNG" in factory_value:
        return "factory"

    if "BẾN" in station_value:
        return "station"

    return "other"


class Command(BaseCommand):
    help = "Import dữ liệu vận chuyển từ file Excel"

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str)

    def handle(self, *args, **options):
        excel_path = options["excel_path"]

        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active

        self.stdout.write(self.style.WARNING("Bắt đầu import dữ liệu vận chuyển..."))

        company_count = 0
        route_count = 0
        relation_count = 0

        # PHẦN 1: IMPORT DANH MỤC NHÀ XE
        # Dữ liệu bắt đầu từ dòng 3
        for row in range(3, ws.max_row + 1):
            company_name = clean_value(ws[f"I{row}"].value)

            if not company_name:
                continue

            phone = clean_value(ws[f"J{row}"].value)
            receive_time = clean_value(ws[f"K{row}"].value)
            departure_time = clean_value(ws[f"L{row}"].value)
            factory_receive = clean_value(ws[f"M{row}"].value)
            station_receive = clean_value(ws[f"N{row}"].value)

            receive_type = detect_receive_type(factory_receive, station_receive)

            company, created = TransportCompany.objects.update_or_create(
                name=company_name,
                defaults={
                    "phone": phone,
                    "receive_time": receive_time,
                    "departure_time": departure_time,
                    "receive_type": receive_type,
                    "is_active": True,
                }
            )

            if created:
                company_count += 1

        # PHẦN 2: IMPORT TUYẾN VẬN CHUYỂN + GÁN NHÀ XE
        for row in range(3, ws.max_row + 1):
            route_name = clean_value(ws[f"B{row}"].value)

            if not route_name or route_name == "-":
                continue

            route, created = TransportRoute.objects.update_or_create(
                name=route_name,
                defaults={
                    "is_active": True,
                }
            )

            if created:
                route_count += 1

            # Cột C-G tương ứng ưu tiên 1-5
            company_columns = ["C", "D", "E", "F", "G"]

            for index, col in enumerate(company_columns, start=1):
                company_name = clean_value(ws[f"{col}{row}"].value)

                if not company_name:
                    continue

                company, _ = TransportCompany.objects.get_or_create(
                    name=company_name,
                    defaults={
                        "receive_type": "other",
                        "is_active": True,
                    }
                )

                relation, created = RouteCompany.objects.update_or_create(
                    route=route,
                    company=company,
                    defaults={
                        "priority": index,
                    }
                )

                if created:
                    relation_count += 1

        self.stdout.write(self.style.SUCCESS("Import hoàn tất!"))
        self.stdout.write(f"Nhà xe mới: {company_count}")
        self.stdout.write(f"Tuyến mới: {route_count}")
        self.stdout.write(f"Liên kết tuyến - nhà xe mới: {relation_count}")