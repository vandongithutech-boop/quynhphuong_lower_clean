from django.shortcuts import render
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q, Sum
from processing.models import ProcessingTicketItem
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from inventory.models import (
    InventoryStock,
    InventoryTransaction,
)
from purchases.models import PurchaseOrder


def get_date_range(request):
    from_date_str = request.GET.get("from_date", "")
    to_date_str = request.GET.get("to_date", "")

    today = timezone.localdate()

    try:
        from_date = (
            timezone.datetime.strptime(from_date_str, "%Y-%m-%d").date()
            if from_date_str else today
        )
    except ValueError:
        from_date = today

    try:
        to_date = (
            timezone.datetime.strptime(to_date_str, "%Y-%m-%d").date()
            if to_date_str else from_date
        )
    except ValueError:
        to_date = from_date

    current_tz = timezone.get_current_timezone()

    start = timezone.make_aware(
        timezone.datetime.combine(from_date, timezone.datetime.min.time()),
        current_tz
    )

    end = timezone.make_aware(
        timezone.datetime.combine(to_date, timezone.datetime.max.time()),
        current_tz
    )

    return from_date, to_date, start, end


def report_dashboard(request):
    return render(request, "reports/dashboard.html")


def build_inventory_current_data(request):
    from_date, to_date, start, end = get_date_range(request)

    q = request.GET.get("q", "").strip()
    warehouse = request.GET.get("warehouse", "")
    group = request.GET.get("group", "")

    before_transactions = InventoryTransaction.objects.filter(
        business_date__lt=from_date
    )

    period_transactions = InventoryTransaction.objects.filter(
        business_date__range=(from_date, to_date)
    )

    if q:
        before_transactions = before_transactions.filter(
            Q(product_code__icontains=q) |
            Q(product_name__icontains=q)
        )

        period_transactions = period_transactions.filter(
            Q(product_code__icontains=q) |
            Q(product_name__icontains=q)
        )

    if warehouse:
        before_transactions = before_transactions.filter(
            warehouse_type=warehouse
        )

        period_transactions = period_transactions.filter(
            warehouse_type=warehouse
        )

    if group:
        before_transactions = before_transactions.filter(
            product_group=group
        )

        period_transactions = period_transactions.filter(
            product_group=group
        )

    report_map = {}

    def clean_text(value):
        return str(value or "").strip()

    def clean_grade(value):
        value = str(value or "").strip().lower()

        if value in ["", "-", "none", "null", "chưa có dữ liệu"]:
            return ""

        return value

    def get_group_label(product_group):
        for value, label in InventoryStock.PRODUCT_GROUP_CHOICES:
            if value == product_group:
                return label

        return product_group or "-"

    def get_warehouse_label(warehouse_type):
        for value, label in InventoryStock.WAREHOUSE_CHOICES:
            if value == warehouse_type:
                return label

        return warehouse_type or "-"

    def make_key(product_code, product_name, product_group, unit, warehouse_type, flower_grade):
        return (
            clean_text(product_code),
            clean_text(product_name),
            clean_text(product_group),
            clean_text(unit),
            clean_text(warehouse_type),
            clean_grade(flower_grade),
        )

    def ensure_row(product_code, product_name, product_group, unit, warehouse_type, flower_grade):
        key = make_key(
            product_code,
            product_name,
            product_group,
            unit,
            warehouse_type,
            flower_grade,
        )

        if key not in report_map:
            clean_flower_grade = clean_grade(flower_grade)

            report_map[key] = {
                "product_code": clean_text(product_code) or "-",
                "product_name": clean_text(product_name) or "-",
                "product_group": get_group_label(product_group),
                "unit": clean_text(unit) or "-",

                "warehouse_type": clean_text(warehouse_type) or "-",
                "warehouse_label": get_warehouse_label(warehouse_type),
                "flower_grade": clean_flower_grade or "-",

                "opening_qty": 0,
                "in_qty": 0,
                "out_qty": 0,
                "check_qty": 0,
                "closing_qty": 0,
            }

        return report_map[key]

    def apply_transaction_to_row(row, trans, mode):
        qty = float(trans.quantity or 0)

        if trans.transaction_type == "IN":
            if mode == "before":
                row["opening_qty"] += abs(qty)
            else:
                row["in_qty"] += abs(qty)

        elif trans.transaction_type == "OUT":
            if mode == "before":
                row["opening_qty"] -= abs(qty)
            else:
                row["out_qty"] += abs(qty)

        elif trans.transaction_type == "CHECK":
            if mode == "before":
                row["opening_qty"] += qty
            else:
                row["check_qty"] += qty

                if qty > 0:
                    row["in_qty"] += abs(qty)
                elif qty < 0:
                    row["out_qty"] += abs(qty)

        elif trans.transaction_type == "ADJUST":
            if mode == "before":
                row["opening_qty"] += qty
            else:
                row["check_qty"] += qty

                if qty > 0:
                    row["in_qty"] += abs(qty)
                elif qty < 0:
                    row["out_qty"] += abs(qty)

        elif trans.transaction_type == "TRANSFER":
            # Hiện tại TRANSFER chưa dùng rõ chiều âm/dương.
            # Nếu sau này lưu TRANSFER bằng quantity âm/dương thì mở phần này.
            if mode == "before":
                row["opening_qty"] += qty
            else:
                row["check_qty"] += qty

                if qty > 0:
                    row["in_qty"] += abs(qty)
                elif qty < 0:
                    row["out_qty"] += abs(qty)

    for trans in before_transactions:
        row = ensure_row(
            trans.product_code,
            trans.product_name,
            trans.product_group,
            trans.unit,
            trans.warehouse_type,
            trans.flower_grade,
        )

        apply_transaction_to_row(row, trans, mode="before")

    for trans in period_transactions:
        row = ensure_row(
            trans.product_code,
            trans.product_name,
            trans.product_group,
            trans.unit,
            trans.warehouse_type,
            trans.flower_grade,
        )

        apply_transaction_to_row(row, trans, mode="period")

    for row in report_map.values():
        row["closing_qty"] = (
            row["opening_qty"]
            + row["in_qty"]
            - row["out_qty"]
        )

    report_rows = sorted(
        report_map.values(),
        key=lambda x: (
            x["product_name"],
            x["warehouse_type"],
            x["flower_grade"],
        )
    )

    return {
        "from_date": from_date,
        "to_date": to_date,
        "q": q,
        "warehouse": warehouse,
        "group": group,
        "report_rows": report_rows,

        "total_opening_qty": sum(row["opening_qty"] for row in report_rows),
        "total_in_qty": sum(row["in_qty"] for row in report_rows),
        "total_out_qty": sum(row["out_qty"] for row in report_rows),
        "total_check_qty": sum(row["check_qty"] for row in report_rows),
        "total_closing_qty": sum(row["closing_qty"] for row in report_rows),

        "total_stock_qty": sum(row["closing_qty"] for row in report_rows),
        "total_product_count": len(report_rows),

        "warehouse_choices": InventoryStock.WAREHOUSE_CHOICES,
        "product_group_choices": InventoryStock.PRODUCT_GROUP_CHOICES,
    }


def inventory_current_report(request):
    context = build_inventory_current_data(request)

    today = timezone.localdate()
    yesterday = today - timezone.timedelta(days=1)

    today_start = timezone.make_aware(
        timezone.datetime.combine(today, timezone.datetime.min.time())
    )
    today_end = timezone.make_aware(
        timezone.datetime.combine(today, timezone.datetime.max.time())
    )

    yesterday_start = timezone.make_aware(
        timezone.datetime.combine(yesterday, timezone.datetime.min.time())
    )
    yesterday_end = timezone.make_aware(
        timezone.datetime.combine(yesterday, timezone.datetime.max.time())
    )

    today_in = (
        InventoryTransaction.objects
        .filter(transaction_type="IN", created_at__range=(today_start, today_end))
        .values("product_name")
        .annotate(total_quantity=Sum("quantity"))
        .order_by("-total_quantity")[:12]
    )

    yesterday_in = (
        InventoryTransaction.objects
        .filter(transaction_type="IN", created_at__range=(yesterday_start, yesterday_end))
        .values("product_name")
        .annotate(total_quantity=Sum("quantity"))
    )

    yesterday_map = {
        item["product_name"]: float(item["total_quantity"] or 0)
        for item in yesterday_in
    }

    chart_labels = []
    chart_today = []
    chart_yesterday = []

    for item in today_in:
        name = item["product_name"] or "-"
        chart_labels.append(name)
        chart_today.append(float(item["total_quantity"] or 0))
        chart_yesterday.append(yesterday_map.get(name, 0))

    context.update({
        "today_in_total": sum(chart_today),
        "yesterday_in_total": sum(chart_yesterday),
        "chart_labels": chart_labels,
        "chart_today": chart_today,
        "chart_yesterday": chart_yesterday,
    })

    return render(request, "reports/inventory_current_report.html", context)


def inventory_in_report(request):
    from_date, to_date, start, end = get_date_range(request)

    q = request.GET.get("q", "").strip()
    warehouse = request.GET.get("warehouse", "")
    group = request.GET.get("group", "")

    transactions = InventoryTransaction.objects.filter(
        Q(transaction_type="IN") |
        Q(transaction_type="CHECK", quantity__gt=0),
        business_date__range=(from_date, to_date),
    ).order_by("-business_date", "-created_at")

    if q:
        transactions = transactions.filter(
            Q(product_code__icontains=q) |
            Q(product_name__icontains=q) |
            Q(reference_code__icontains=q)
        )

    if warehouse:
        transactions = transactions.filter(warehouse_type=warehouse)

    if group:
        transactions = transactions.filter(product_group=group)

    top_map = {}

    for trans in transactions:
        key = (
            trans.product_code or "-",
            trans.product_name or "-",
            trans.unit or "-",
        )

        if key not in top_map:
            top_map[key] = {
                "product_code": trans.product_code or "-",
                "product_name": trans.product_name or "-",
                "unit": trans.unit or "-",
                "total_quantity": 0,
            }

        top_map[key]["total_quantity"] += abs(float(trans.quantity or 0))

    top_products = sorted(
        top_map.values(),
        key=lambda x: x["total_quantity"],
        reverse=True
    )[:10]

    context = {
        "from_date": from_date,
        "to_date": to_date,
        "q": q,
        "warehouse": warehouse,
        "group": group,
        "transactions": transactions,
        "total_quantity": transactions.aggregate(total=Sum("quantity"))["total"] or 0,
        "total_transaction_count": transactions.count(),
        "top_products": top_products,
        "warehouse_choices": InventoryStock.WAREHOUSE_CHOICES,
        "product_group_choices": InventoryStock.PRODUCT_GROUP_CHOICES,
    }

    return render(request, "reports/inventory_in_report.html", context)


def inventory_out_report(request):
    from_date, to_date, start, end = get_date_range(request)

    q = request.GET.get("q", "").strip()
    warehouse = request.GET.get("warehouse", "")
    group = request.GET.get("group", "")

    transactions = InventoryTransaction.objects.filter(
        Q(transaction_type="OUT") |
        Q(transaction_type="CHECK", quantity__lt=0),
        business_date__range=(from_date, to_date),
    ).order_by("-business_date", "-created_at")

    if q:
        transactions = transactions.filter(
            Q(product_code__icontains=q) |
            Q(product_name__icontains=q) |
            Q(reference_code__icontains=q)
        )

    if warehouse:
        transactions = transactions.filter(warehouse_type=warehouse)

    if group:
        transactions = transactions.filter(product_group=group)

    top_products = (
        transactions
        .values("product_code", "product_name", "unit")
        .annotate(total_quantity=Sum("quantity"))
        .order_by("-total_quantity")[:10]
    )

    total_quantity = 0
    for item in transactions:
        total_quantity += abs(float(item.quantity or 0))

    context = {
        "from_date": from_date,
        "to_date": to_date,
        "q": q,
        "warehouse": warehouse,
        "group": group,
        "transactions": transactions,
        "total_quantity": total_quantity,
        "total_transaction_count": transactions.count(),
        "top_products": top_products,
        "warehouse_choices": InventoryStock.WAREHOUSE_CHOICES,
        "product_group_choices": InventoryStock.PRODUCT_GROUP_CHOICES,
    }

    return render(request, "reports/inventory_out_report.html", context)


def export_inventory_report_excel(request):
    context = build_inventory_current_data(request)
    report_rows = context["report_rows"]

    from_date = context["from_date"]
    to_date = context["to_date"]
    q = context.get("q", "")
    warehouse = context.get("warehouse", "")
    group = context.get("group", "")

    def get_product_group_label(product_group):
        for value, label in InventoryStock.PRODUCT_GROUP_CHOICES:
            if value == product_group:
                return label
        return product_group or "-"

    def get_warehouse_label(warehouse_type):
        for value, label in InventoryStock.WAREHOUSE_CHOICES:
            if value == warehouse_type:
                return label
        return warehouse_type or "-"

    in_transactions = InventoryTransaction.objects.filter(
        Q(transaction_type="IN") |
        Q(transaction_type="CHECK", quantity__gt=0),
        business_date__range=(from_date, to_date),
    ).order_by("business_date", "created_at")

    if q:
        in_transactions = in_transactions.filter(
            Q(product_code__icontains=q) |
            Q(product_name__icontains=q) |
            Q(reference_code__icontains=q)
        )

    if warehouse:
        in_transactions = in_transactions.filter(warehouse_type=warehouse)

    if group:
        in_transactions = in_transactions.filter(product_group=group)

    processing_items = (
        ProcessingTicketItem.objects
        .select_related("ticket")
        .filter(ticket__business_date__range=(from_date, to_date))
        .order_by("ticket__business_date", "ticket__ticket_code", "id")
    )

    if q:
        processing_items = processing_items.filter(
            Q(product_code__icontains=q) |
            Q(product_name__icontains=q) |
            Q(ticket__ticket_code__icontains=q) |
            Q(ticket__employee_code__icontains=q) |
            Q(ticket__employee_name__icontains=q)
        )

    if group:
        processing_items = processing_items.filter(product_group=group)

    wb = Workbook()

    green_fill = PatternFill("solid", fgColor="15803D")
    dark_green_fill = PatternFill("solid", fgColor="14532D")
    light_green_fill = PatternFill("solid", fgColor="DCFCE7")
    pale_green_fill = PatternFill("solid", fgColor="ECFDF5")
    yellow_fill = PatternFill("solid", fgColor="FEF3C7")
    blue_fill = PatternFill("solid", fgColor="DBEAFE")
    red_fill = PatternFill("solid", fgColor="FEE2E2")
    orange_fill = PatternFill("solid", fgColor="FFEDD5")
    gray_fill = PatternFill("solid", fgColor="F8FAFC")
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=18, bold=True, color="14532D")
    subtitle_font = Font(size=11, bold=True, color="166534")
    section_font = Font(size=13, bold=True, color="14532D")
    bold_font = Font(bold=True)
    small_gray_font = Font(size=10, color="64748B")

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # =========================================================
    # SHEET 1: BÁO CÁO TỒN KHO
    # =========================================================

    ws = wb.active
    ws.title = "Bao cao ton kho"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    ws.merge_cells("A1:N1")
    ws["A1"] = "CÔNG TY TNHH QUỲNH PHƯƠNG ĐÀ LẠT"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.merge_cells("A2:N2")
    ws["A2"] = "QUỲNH PHƯƠNG FLOWER EXPORT SYSTEM"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = center

    ws.merge_cells("A4:N4")
    ws["A4"] = "BÁO CÁO TỒN KHO CHI TIẾT THEO SẢN PHẨM"
    ws["A4"].font = Font(size=15, bold=True, color="0F172A")
    ws["A4"].alignment = center

    ws.merge_cells("A5:N5")
    ws["A5"] = (
        f"Từ ngày {from_date.strftime('%d/%m/%Y')} "
        f"đến ngày {to_date.strftime('%d/%m/%Y')}"
    )
    ws["A5"].font = small_gray_font
    ws["A5"].alignment = center

    summary_start = 7
    summary_data = [
        ("Tồn đầu kỳ", context["total_opening_qty"]),
        ("Nhập trong kỳ", context["total_in_qty"]),
        ("Xuất trong kỳ", context["total_out_qty"]),
        ("Kiểm kê", context["total_check_qty"]),
        ("Tồn cuối kỳ", context["total_closing_qty"]),
        ("Số dòng", len(report_rows)),
    ]

    for i, (label, value) in enumerate(summary_data):
        col = 1 + i * 2

        ws.merge_cells(
            start_row=summary_start,
            start_column=col,
            end_row=summary_start,
            end_column=col + 1
        )

        label_cell = ws.cell(row=summary_start, column=col, value=label)
        label_cell.fill = dark_green_fill
        label_cell.font = white_font
        label_cell.alignment = center
        label_cell.border = border

        ws.merge_cells(
            start_row=summary_start + 1,
            start_column=col,
            end_row=summary_start + 1,
            end_column=col + 1
        )

        value_cell = ws.cell(row=summary_start + 1, column=col, value=value)
        value_cell.fill = pale_green_fill
        value_cell.font = Font(size=13, bold=True, color="14532D")
        value_cell.alignment = center
        value_cell.border = border

    ws.merge_cells("A11:N11")
    ws["A11"] = "BẢNG TỒN KHO CHI TIẾT THEO SẢN PHẨM"
    ws["A11"].font = section_font
    ws["A11"].alignment = left

    headers = [
        "STT",
        "Mã SP",
        "Tên sản phẩm",
        "Nhóm hàng",
        "Kho",
        "Loại",
        "ĐVT",
        "Tồn đầu",
        "Nhập",
        "Xuất",
        "Kiểm kê",
        "Sale",
        "Hủy",
        "Tồn cuối",
    ]

    start_row = 13

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = wrap_center
        cell.border = border

    grouped_sections = [
        (
            "PHẦN 1: HOA HỒNG",
            [row for row in report_rows if row.get("product_group") == "Hoa Hồng"]
        ),
        (
            "PHẦN 2: HOA/LÁ PHỤ",
            [row for row in report_rows if row.get("product_group") == "Hoa/Lá Phụ"]
        ),
        (
            "PHẦN 3: VẬT TƯ / KHÁC",
            [
                row for row in report_rows
                if row.get("product_group") not in ["Hoa Hồng", "Hoa/Lá Phụ"]
            ]
        ),
    ]

    current_row = start_row + 1
    index = 1

    for section_title, section_rows in grouped_sections:
        if not section_rows:
            continue

        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=14
        )

        section_cell = ws.cell(row=current_row, column=1, value=section_title)
        section_cell.fill = dark_green_fill
        section_cell.font = white_font
        section_cell.alignment = left
        section_cell.border = border

        for col in range(1, 15):
            ws.cell(row=current_row, column=col).border = border

        current_row += 1

        for row in section_rows:
            is_sale = row.get("warehouse_type") == "SALE"
            is_damaged = row.get("warehouse_type") == "DAMAGED"

            sale_qty = row["closing_qty"] if is_sale else 0
            damaged_qty = row["closing_qty"] if is_damaged else 0

            values = [
                index,
                row["product_code"],
                row["product_name"],
                row["product_group"],
                row["warehouse_label"],
                row["flower_grade"],
                row["unit"],
                row["opening_qty"],
                row["in_qty"],
                row["out_qty"],
                row["check_qty"],
                sale_qty,
                damaged_qty,
                row["closing_qty"],
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = border
                cell.alignment = right if col >= 8 else wrap_left
                cell.fill = gray_fill if index % 2 == 0 else white_fill

                if col == 8:
                    cell.fill = blue_fill
                elif col == 9:
                    cell.fill = light_green_fill
                elif col == 10:
                    cell.fill = red_fill
                elif col == 11:
                    cell.fill = orange_fill
                elif col == 12:
                    cell.fill = pale_green_fill
                elif col == 13:
                    cell.fill = red_fill
                elif col == 14:
                    cell.fill = light_green_fill
                    cell.font = bold_font

            current_row += 1
            index += 1

        current_row += 1

    total_row = current_row

    ws.cell(row=total_row, column=1, value="TỔNG CỘNG")
    ws.merge_cells(
        start_row=total_row,
        start_column=1,
        end_row=total_row,
        end_column=7
    )

    for col in range(1, 15):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = yellow_fill
        cell.font = bold_font
        cell.border = border
        cell.alignment = center if col <= 7 else right

    for col in range(8, 15):
        col_letter = get_column_letter(col)
        ws.cell(
            row=total_row,
            column=col,
            value=f"=SUM({col_letter}{start_row + 1}:{col_letter}{total_row - 1})"
        )

    signature_row = total_row + 3

    ws.merge_cells(
        start_row=signature_row,
        start_column=2,
        end_row=signature_row,
        end_column=4
    )
    ws.cell(signature_row, 2, "Người lập báo cáo").font = bold_font
    ws.cell(signature_row, 2).alignment = center

    ws.merge_cells(
        start_row=signature_row,
        start_column=6,
        end_row=signature_row,
        end_column=8
    )
    ws.cell(signature_row, 6, "Thủ kho").font = bold_font
    ws.cell(signature_row, 6).alignment = center

    ws.merge_cells(
        start_row=signature_row,
        start_column=10,
        end_row=signature_row,
        end_column=12
    )
    ws.cell(signature_row, 10, "Ban giám đốc").font = bold_font
    ws.cell(signature_row, 10).alignment = center

    widths = {
        "A": 8,
        "B": 16,
        "C": 38,
        "D": 18,
        "E": 22,
        "F": 12,
        "G": 12,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 14,
        "L": 14,
        "M": 14,
        "N": 14,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row_num in range(1, total_row + 8):
        ws.row_dimensions[row_num].height = 24

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[start_row].height = 36

    ws.freeze_panes = "A14"
    ws.auto_filter.ref = f"A{start_row}:N{total_row}"
    ws.print_title_rows = f"{start_row}:{start_row}"
    ws.oddFooter.center.text = "QUỲNH PHƯƠNG FLOWER EXPORT SYSTEM"
    ws.oddFooter.right.text = "Trang &[Page]/&[Pages]"

    # =========================================================
    # SHEET 2: BÁO CÁO NHẬP KHO
    # =========================================================

    ws_in = wb.create_sheet("Bao cao nhap kho")
    ws_in.sheet_view.showGridLines = False
    ws_in.page_setup.orientation = "landscape"
    ws_in.page_setup.paperSize = ws_in.PAPERSIZE_A4
    ws_in.page_margins.left = 0.25
    ws_in.page_margins.right = 0.25
    ws_in.page_margins.top = 0.5
    ws_in.page_margins.bottom = 0.5

    ws_in.merge_cells("A1:L1")
    ws_in["A1"] = "CÔNG TY TNHH QUỲNH PHƯƠNG ĐÀ LẠT"
    ws_in["A1"].font = title_font
    ws_in["A1"].alignment = center

    ws_in.merge_cells("A2:L2")
    ws_in["A2"] = "QUỲNH PHƯƠNG FLOWER EXPORT SYSTEM"
    ws_in["A2"].font = subtitle_font
    ws_in["A2"].alignment = center

    ws_in.merge_cells("A4:L4")
    ws_in["A4"] = "BÁO CÁO NHẬP KHO"
    ws_in["A4"].font = Font(size=15, bold=True, color="0F172A")
    ws_in["A4"].alignment = center

    ws_in.merge_cells("A5:L5")
    ws_in["A5"] = (
        f"Từ ngày {from_date.strftime('%d/%m/%Y')} "
        f"đến ngày {to_date.strftime('%d/%m/%Y')}"
    )
    ws_in["A5"].font = small_gray_font
    ws_in["A5"].alignment = center

    in_headers = [
        "STT",
        "Ngày nghiệp vụ",
        "Mã phiếu",
        "Kho",
        "Nhóm",
        "Mã SP",
        "Tên sản phẩm",
        "ĐVT",
        "Độ / loại",
        "Số lượng",
        "Nhà cung cấp",
        "Ghi chú",
    ]

    in_start_row = 7

    for col, header in enumerate(in_headers, 1):
        cell = ws_in.cell(row=in_start_row, column=col, value=header)
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = wrap_center
        cell.border = border

    in_row = in_start_row + 1

    for idx, trans in enumerate(in_transactions, 1):
        values = [
            idx,
            trans.business_date.strftime("%d/%m/%Y") if trans.business_date else "",
            trans.reference_code or "-",
            get_warehouse_label(trans.warehouse_type),
            get_product_group_label(trans.product_group),
            trans.product_code or "-",
            trans.product_name or "-",
            trans.unit or "-",
            trans.flower_grade or "-",
            abs(float(trans.quantity or 0)),
            trans.supplier_name or "-",
            trans.note or "-",
        ]

        for col, value in enumerate(values, 1):
            cell = ws_in.cell(row=in_row, column=col, value=value)
            cell.border = border
            cell.alignment = right if col == 10 else wrap_left
            cell.fill = gray_fill if idx % 2 == 0 else white_fill

            if col == 10:
                cell.fill = light_green_fill
                cell.font = bold_font

        in_row += 1

    total_in_row = in_row

    ws_in.cell(row=total_in_row, column=1, value="TỔNG CỘNG")
    ws_in.merge_cells(
        start_row=total_in_row,
        start_column=1,
        end_row=total_in_row,
        end_column=9
    )

    for col in range(1, 13):
        cell = ws_in.cell(row=total_in_row, column=col)
        cell.fill = yellow_fill
        cell.font = bold_font
        cell.border = border
        cell.alignment = center if col <= 9 else right

    ws_in.cell(
        row=total_in_row,
        column=10,
        value=f"=SUM(J{in_start_row + 1}:J{total_in_row - 1})"
    )

    in_widths = {
        "A": 8,
        "B": 16,
        "C": 22,
        "D": 22,
        "E": 16,
        "F": 16,
        "G": 38,
        "H": 12,
        "I": 14,
        "J": 14,
        "K": 28,
        "L": 42,
    }

    for col, width in in_widths.items():
        ws_in.column_dimensions[col].width = width

    for row_num in range(1, total_in_row + 2):
        ws_in.row_dimensions[row_num].height = 24

    ws_in.row_dimensions[1].height = 30
    ws_in.row_dimensions[4].height = 28
    ws_in.row_dimensions[in_start_row].height = 36

    ws_in.freeze_panes = "A8"
    ws_in.auto_filter.ref = f"A{in_start_row}:L{total_in_row}"
    ws_in.print_title_rows = f"{in_start_row}:{in_start_row}"
    ws_in.oddFooter.center.text = "QUỲNH PHƯƠNG FLOWER EXPORT SYSTEM"
    ws_in.oddFooter.right.text = "Trang &[Page]/&[Pages]"

    # =========================================================
    # SHEET 3: BÁO CÁO SƠ CHẾ
    # =========================================================

    ws_processing = wb.create_sheet("Bao cao so che")
    ws_processing.sheet_view.showGridLines = False
    ws_processing.page_setup.orientation = "landscape"
    ws_processing.page_setup.paperSize = ws_processing.PAPERSIZE_A4
    ws_processing.page_margins.left = 0.25
    ws_processing.page_margins.right = 0.25
    ws_processing.page_margins.top = 0.5
    ws_processing.page_margins.bottom = 0.5

    ws_processing.merge_cells("A1:AD1")
    ws_processing["A1"] = "CÔNG TY TNHH QUỲNH PHƯƠNG ĐÀ LẠT"
    ws_processing["A1"].font = title_font
    ws_processing["A1"].alignment = center

    ws_processing.merge_cells("A2:AD2")
    ws_processing["A2"] = "QUỲNH PHƯƠNG FLOWER EXPORT SYSTEM"
    ws_processing["A2"].font = subtitle_font
    ws_processing["A2"].alignment = center

    ws_processing.merge_cells("A4:AD4")
    ws_processing["A4"] = "BÁO CÁO SƠ CHẾ CHI TIẾT"
    ws_processing["A4"].font = Font(size=15, bold=True, color="0F172A")
    ws_processing["A4"].alignment = center

    ws_processing.merge_cells("A5:AD5")
    ws_processing["A5"] = (
        f"Từ ngày {from_date.strftime('%d/%m/%Y')} "
        f"đến ngày {to_date.strftime('%d/%m/%Y')}"
    )
    ws_processing["A5"].font = small_gray_font
    ws_processing["A5"].alignment = center

    processing_headers = [
        "STT",
        "Ngày nghiệp vụ",
        "Mã phiếu",
        "Mã NV",
        "Nhân viên",
        "Chức vụ",
        "Ca 1",
        "Giờ ca 1",
        "Ca 2",
        "Giờ ca 2",
        "Tổng giờ",
        "Tăng ca",
        "Mã SP",
        "Tên sản phẩm",
        "Nhóm",
        "ĐVT",
        "Độ hoa",
        "Nhà cung cấp",
        "Ngày nhập gốc",
        "Tồn lúc sơ chế",
        "SL nhận",
        "Quy cách",
        "Cành sơ chế",
        "Xả/Hủy",
        "Cành lẻ",
        "Bù cành lẻ khác",
        "Cành thừa",
        "Cành TP",
        "Bó TP",
        "Ghi chú",
    ]

    processing_start_row = 7

    for col, header in enumerate(processing_headers, 1):
        cell = ws_processing.cell(row=processing_start_row, column=col, value=header)
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = wrap_center
        cell.border = border

    processing_row = processing_start_row + 1

    def format_time_range(start_value, end_value):
        if start_value and end_value:
            return f"{start_value.strftime('%H:%M')} - {end_value.strftime('%H:%M')}"
        return "-"

    for idx, item in enumerate(processing_items, 1):
        ticket = item.ticket
        compensate_qty = float(getattr(item, "compensate_stems", 0) or 0)

        received_date_value = ""
        if item.received_date:
            received_date_value = item.received_date.strftime("%d/%m/%Y")

        values = [
            idx,
            ticket.business_date.strftime("%d/%m/%Y") if ticket.business_date else "",
            ticket.ticket_code or "-",
            ticket.employee_code or "-",
            ticket.employee_name or "-",
            ticket.employee_position or "-",
            format_time_range(ticket.shift1_start, ticket.shift1_end),
            float(ticket.shift1_hours or 0),
            format_time_range(ticket.shift2_start, ticket.shift2_end),
            float(ticket.shift2_hours or 0),
            float(ticket.total_work_hours or ticket.total_hours or 0),
            float(ticket.overtime_hours or 0),
            item.product_code or "-",
            item.product_name or "-",
            get_product_group_label(item.product_group),
            item.unit or "-",
            item.flower_grade or "-",
            item.supplier_name or "-",
            received_date_value,
            float(item.stock_quantity or 0),
            float(item.received_quantity or 0),
            item.bunch_type or "-",
            float(item.processed_stems or 0),
            float(item.damaged_stems or 0),
            float(item.odd_stems or 0),
            compensate_qty,
            float(item.extra_stems or 0),
            float(item.final_stems or 0),
            float(item.final_bunches or 0),
            ticket.work_note or "-",
        ]

        for col, value in enumerate(values, 1):
            cell = ws_processing.cell(row=processing_row, column=col, value=value)
            cell.border = border
            cell.alignment = right if col in [8, 10, 11, 12, 20, 21, 23, 24, 25, 26, 27, 28, 29] else wrap_left
            cell.fill = gray_fill if idx % 2 == 0 else white_fill

            if col in [21, 23, 28, 29]:
                cell.fill = light_green_fill
                cell.font = bold_font
            elif col == 24:
                cell.fill = red_fill
            elif col == 25:
                cell.fill = orange_fill
            elif col == 26:
                cell.fill = pale_green_fill
            elif col == 27:
                cell.fill = blue_fill

        processing_row += 1

    total_processing_row = processing_row

    ws_processing.cell(row=total_processing_row, column=1, value="TỔNG CỘNG")
    ws_processing.merge_cells(
        start_row=total_processing_row,
        start_column=1,
        end_row=total_processing_row,
        end_column=19
    )

    for col in range(1, 31):
        cell = ws_processing.cell(row=total_processing_row, column=col)
        cell.fill = yellow_fill
        cell.font = bold_font
        cell.border = border
        cell.alignment = center if col <= 19 else right

    ws_processing.cell(row=total_processing_row, column=20, value=f"=SUM(T{processing_start_row + 1}:T{total_processing_row - 1})")
    ws_processing.cell(row=total_processing_row, column=21, value=f"=SUM(U{processing_start_row + 1}:U{total_processing_row - 1})")
    ws_processing.cell(row=total_processing_row, column=23, value=f"=SUM(W{processing_start_row + 1}:W{total_processing_row - 1})")
    ws_processing.cell(row=total_processing_row, column=24, value=f"=SUM(X{processing_start_row + 1}:X{total_processing_row - 1})")
    ws_processing.cell(row=total_processing_row, column=25, value=f"=SUM(Y{processing_start_row + 1}:Y{total_processing_row - 1})")
    ws_processing.cell(row=total_processing_row, column=26, value=f"=SUM(Z{processing_start_row + 1}:Z{total_processing_row - 1})")
    ws_processing.cell(row=total_processing_row, column=27, value=f"=SUM(AA{processing_start_row + 1}:AA{total_processing_row - 1})")
    ws_processing.cell(row=total_processing_row, column=28, value=f"=SUM(AB{processing_start_row + 1}:AB{total_processing_row - 1})")
    ws_processing.cell(row=total_processing_row, column=29, value=f"=SUM(AC{processing_start_row + 1}:AC{total_processing_row - 1})")

    processing_widths = {
        "A": 8,
        "B": 16,
        "C": 22,
        "D": 14,
        "E": 24,
        "F": 18,
        "G": 18,
        "H": 12,
        "I": 18,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 16,
        "N": 38,
        "O": 16,
        "P": 12,
        "Q": 12,
        "R": 28,
        "S": 16,
        "T": 16,
        "U": 14,
        "V": 14,
        "W": 14,
        "X": 14,
        "Y": 14,
        "Z": 16,
        "AA": 14,
        "AB": 14,
        "AC": 14,
        "AD": 36,
    }

    for col, width in processing_widths.items():
        ws_processing.column_dimensions[col].width = width

    for row_num in range(1, total_processing_row + 2):
        ws_processing.row_dimensions[row_num].height = 24

    ws_processing.row_dimensions[1].height = 30
    ws_processing.row_dimensions[4].height = 28
    ws_processing.row_dimensions[processing_start_row].height = 38

    ws_processing.freeze_panes = "A8"
    ws_processing.auto_filter.ref = f"A{processing_start_row}:AD{total_processing_row}"
    ws_processing.print_title_rows = f"{processing_start_row}:{processing_start_row}"
    ws_processing.oddFooter.center.text = "QUỲNH PHƯƠNG FLOWER EXPORT SYSTEM"
    ws_processing.oddFooter.right.text = "Trang &[Page]/&[Pages]"

    file_name = (
        f"bao_cao_ton_kho_"
        f"{from_date.strftime('%Y%m%d')}_"
        f"{to_date.strftime('%Y%m%d')}.xlsx"
    )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'

    wb.save(response)
    return response