import json

from django.shortcuts import render, redirect
from django.utils import timezone
from django.contrib import messages
from django.db.models import Q
from .models import InventoryReceiptItem
from collections import defaultdict
from categories.models import FlowerType, Customer
from traceability.services import create_trace_lot_from_receipt_item
from django.http import HttpResponse
from django.http import JsonResponse
from purchases.models import PurchaseOrder
from django.db.models import Sum, Count

from django.views.decorators.http import require_POST
from purchases.models import PurchaseOrderItem, PurchaseWarehouseCheckDraft


from django.http import HttpResponse


from django.db import transaction
from traceability.services import create_trace_bundles_from_purchase_item



from purchases.models import PurchaseOrder

from .models import (
    InventoryReceipt,
    InventoryReceiptItem,
    InventoryStock,
    InventoryTransaction,
    InventoryStockOut,
    InventoryStockOutItem,
)

def normalize_flower_grade(value):
    value = str(value or "").strip().lower()

    if value in ["", "-", "none", "null", "chưa có dữ liệu"]:
        return ""

    return value


def get_or_create_normalized_stock(
    warehouse_type,
    product_group,
    product_code,
    product_name,
    unit="-",
    flower_grade="",
    lot=None,
):
    flower_grade = normalize_flower_grade(flower_grade)

    stocks = (
        InventoryStock.objects
        .select_for_update()
        .filter(
            warehouse_type=warehouse_type,
            product_group=product_group,
            product_code=product_code or product_name,
            flower_grade=flower_grade,
        )
        .order_by("id")
    )

    stock = stocks.first()

    if not stock:
        stock = InventoryStock.objects.create(
            lot=lot,
            warehouse_type=warehouse_type,
            product_group=product_group,
            product_code=product_code or product_name,
            product_name=product_name or "-",
            unit=unit or "-",
            flower_grade=flower_grade,
            supplier_code="",
            supplier_name="-",
            received_date=None,
            quantity=0,
        )

        return stock

    duplicate_ids = list(
        stocks.exclude(id=stock.id).values_list("id", flat=True)
    )

    if duplicate_ids:
        total_quantity = sum(float(s.quantity or 0) for s in stocks)

        InventoryStock.objects.filter(id__in=duplicate_ids).delete()

        stock.quantity = total_quantity

    stock.product_name = product_name or stock.product_name or "-"
    stock.unit = unit or stock.unit or "-"
    stock.flower_grade = flower_grade
    stock.supplier_code = ""
    stock.supplier_name = "-"
    stock.received_date = None

    if lot and not stock.lot:
        stock.lot = lot

    stock.save()

    return stock

def get_default_business_date():
    now = timezone.localtime(timezone.now())

    if now.hour < 6:
        return now.date() - timezone.timedelta(days=1)

    return now.date()


def parse_business_date(value):
    if not value:
        return get_default_business_date()

    try:
        return timezone.datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return get_default_business_date()

def inventory_dashboard(request):
    suppliers = Customer.objects.filter(ma_kh__startswith='NCC')

    products_hh = FlowerType.objects.filter(category_type='HH')
    products_hp = FlowerType.objects.filter(category_type='HP')
    products_vt = FlowerType.objects.exclude(category_type__in=['HH', 'HP'])

    current_employee = None

    if request.user.is_authenticated:
        user_employee = getattr(request.user, "employee_profile", None)
        if user_employee and user_employee.status == "approved":
            current_employee = user_employee.employee

    rose_total = InventoryStock.objects.filter(
        product_group="HH",
        warehouse_type__in=["RAW", "FINISHED"],
        quantity__gt=0
    ).aggregate(total=Sum("quantity"))["total"] or 0

    flower_leaf_total = InventoryStock.objects.filter(
        product_group="HP",
        warehouse_type__in=["RAW", "FINISHED"],
        quantity__gt=0
    ).aggregate(total=Sum("quantity"))["total"] or 0

    sale_total = InventoryStock.objects.filter(
        warehouse_type="SALE",
        quantity__gt=0
    ).aggregate(total=Sum("quantity"))["total"] or 0

    material_total = InventoryStock.objects.filter(
        warehouse_type="MATERIAL",
        quantity__gt=0
    ).aggregate(total=Sum("quantity"))["total"] or 0

    damaged_total = InventoryStock.objects.filter(
        warehouse_type="DAMAGED",
        quantity__gt=0
    ).aggregate(total=Sum("quantity"))["total"] or 0
    pending_stock_in_orders = (
        PurchaseOrder.objects
        .filter(status="waiting_stock_in")
        .prefetch_related("items")
        .order_by("-created_at")
    )
    for po in pending_stock_in_orders:
        supplier_names = set()

        for item in po.items.all():
            if item.supplier_name:
                supplier_names.add(item.supplier_name.strip())
            elif po.supplier_name:
                supplier_names.add(po.supplier_name.strip())

        po.transport_supplier_display = (
            "Đơn hàng hỗn hợp" if len(supplier_names) >= 2 else po.supplier_name
        )

    label_created_receipts = (
        InventoryReceipt.objects
        .filter(status="DRAFT_LABEL_CREATED")
        .prefetch_related("items__lot__bundles")
        .order_by("-receipt_datetime")
    )
    context = {
        'suppliers': suppliers,
        'products_hh': products_hh,
        'products_hp': products_hp,
        'products_vt': products_vt,
        'now': timezone.now(),
        'current_employee': current_employee,
        'rose_total': rose_total,
        'flower_leaf_total': flower_leaf_total,
        'sale_total': sale_total,
        'material_total': material_total,
        'damaged_total': damaged_total,
        'pending_stock_in_orders': pending_stock_in_orders,
        "label_created_receipts": label_created_receipts,
        "business_date": get_default_business_date(),
    }

    return render(request, 'inventory/dashboard.html', context)

def create_stock_in(request):
    if request.method == 'POST':
        product_group = request.POST.get('product_group')
        supplier_value = request.POST.get('supplier')
        items_json = request.POST.get('items_json')
        business_date = parse_business_date(request.POST.get("business_date"))

        if not product_group:
            messages.error(request, 'Vui lòng chọn nhóm hàng nhập kho.')
            return redirect('inventory:dashboard')

        if not supplier_value:
            messages.error(request, 'Vui lòng chọn nhà cung cấp.')
            return redirect('inventory:dashboard')

        if not items_json:
            messages.error(request, 'Vui lòng thêm ít nhất một sản phẩm.')
            return redirect('inventory:dashboard')

        try:
            items = json.loads(items_json)
        except json.JSONDecodeError:
            messages.error(request, 'Dữ liệu sản phẩm không hợp lệ.')
            return redirect('inventory:dashboard')

        if not items:
            messages.error(request, 'Vui lòng thêm ít nhất một sản phẩm.')
            return redirect('inventory:dashboard')

        supplier_code = ''
        supplier_name = supplier_value

        if ' - ' in supplier_value:
            supplier_code, supplier_name = supplier_value.split(' - ', 1)

        receipt_code = f"NK{timezone.now().strftime('%Y%m%d%H%M%S')}"

        receipt = InventoryReceipt.objects.create(
            receipt_code=receipt_code,
            product_group=product_group,
            supplier_code=supplier_code,
            supplier_name=supplier_name,
        )

        warehouse_type = 'MATERIAL' if product_group == 'VT' else 'RAW'

        for item in items:
            product_code = item.get('product_code', '')
            product_name = item.get('product_name', '')
            unit = item.get('unit', '')
            flower_grade = item.get('flower_grade', '')
            quantity = float(item.get('quantity') or 0)

            receipt_item = InventoryReceiptItem.objects.create(
                receipt=receipt,
                product_code=product_code,
                product_name=product_name,
                unit=unit,
                flower_grade=flower_grade,
                quantity=quantity,
            )

            lot = create_trace_lot_from_receipt_item(
                receipt,
                receipt_item,
                request=request
            )

            stock, created = InventoryStock.objects.get_or_create(
                warehouse_type='RAW',
                product_group=product_group,
                product_code=item['product_code'],
                flower_grade=item.get('flower_grade') or '',
                supplier_code=receipt.supplier_code,
                received_date=receipt.receipt_datetime,
                defaults={
                    'lot': lot,
                    'product_name': item['product_name'],
                    'unit': item.get('unit'),
                    'supplier_name': receipt.supplier_name,
                    'quantity': 0,
                }
            )

            stock.quantity += float(item['quantity'])
            stock.supplier_name = receipt.supplier_name
            if not stock.lot:
                stock.lot = lot
            stock.save()

            InventoryTransaction.objects.create(
                lot=lot,
                transaction_code=f"GD{timezone.now().strftime('%Y%m%d%H%M%S%f')}",
                transaction_type='IN',
                warehouse_type=warehouse_type,
                product_group=product_group,
                product_code=product_code,
                product_name=product_name,
                unit=unit,
                flower_grade=flower_grade,
                quantity=quantity,
                reference_code=receipt.receipt_code,
                business_date=business_date,
                note='Nhập kho từ phiếu nhập',
            )

        messages.success(
            request,
            f'Đã tạo phiếu nhập kho {receipt_code} thành công.'
        )

        return redirect('inventory:dashboard')

    return redirect('inventory:dashboard')


def grouped_stock_queryset(stocks):
    grouped = {}

    for stock in stocks:
        key = (
            stock.warehouse_type or "",
            stock.product_group or "",
            stock.product_code or "",
            stock.flower_grade or "",
        )

        if key not in grouped:
            stock.grouped_quantity = 0
            stock.grouped_supplier_names = set()
            grouped[key] = stock

        grouped[key].grouped_quantity += float(stock.quantity or 0)

        if stock.supplier_name:
            grouped[key].grouped_supplier_names.add(stock.supplier_name)

    result = []

    for stock in grouped.values():
        stock.quantity = stock.grouped_quantity

        if stock.grouped_supplier_names:
            stock.supplier_name = ", ".join(sorted(stock.grouped_supplier_names))
        else:
            stock.supplier_name = "-"

        result.append(stock)

    return sorted(result, key=lambda x: (x.product_name or "", x.warehouse_type or ""))

def rose_inventory(request):
    q = request.GET.get('q', '')

    stocks = InventoryStock.objects.filter(
        product_group='HH',
        warehouse_type__in=['RAW', 'FINISHED']
    )

    if q:
        stocks = stocks.filter(
            Q(product_code__icontains=q) |
            Q(product_name__icontains=q)
        )

    stocks = grouped_stock_queryset(stocks)

    context = {
        'stocks': stocks,
        'q': q,
        'page_title': 'Kho hoa hồng',
        'page_desc': 'Hiển thị hoa hồng nguyên liệu và hoa hồng thành phẩm',
    }

    return render(request, 'inventory/stock_list.html', context)


def flower_leaf_inventory(request):
    q = request.GET.get('q', '')

    stocks = InventoryStock.objects.filter(
        product_group='HP',
        warehouse_type__in=['RAW', 'FINISHED']
    )

    if q:
        stocks = stocks.filter(
            Q(product_code__icontains=q) |
            Q(product_name__icontains=q)
        )

    stocks = grouped_stock_queryset(stocks)

    context = {
        'stocks': stocks,
        'q': q,
        'page_title': 'Kho hoa/lá phụ',
        'page_desc': 'Hiển thị hoa/lá phụ nguyên liệu và hoa/lá phụ thành phẩm',
    }

    return render(request, 'inventory/stock_list.html', context)



def raw_inventory(request):
    today = timezone.localdate()
    current_filter = request.GET.get("filter", "all")

    stocks = (
        InventoryStock.objects
        .filter(
            warehouse_type="RAW",
            product_group__in=["HH", "HP"],
            quantity__gt=0
        )
        .order_by("received_date", "product_name")
    )

    flower_rows = []

    total_all = 0
    total_new = 0
    total_processing = 0
    total_warning = 0

    for stock in stocks:
        received_date = stock.received_date or stock.updated_at
        received_day = timezone.localtime(received_date).date()
        days_in_stock = (today - received_day).days
        quantity = stock.quantity or 0

        new_qty = quantity if days_in_stock == 0 else 0
        processing_qty = quantity if days_in_stock >= 1 else 0
        warning_qty = quantity if days_in_stock >= 2 else 0

        total_all += quantity
        total_new += new_qty
        total_processing += processing_qty
        total_warning += warning_qty

        show_row = True

        if current_filter == "new":
            show_row = new_qty > 0
        elif current_filter == "process":
            show_row = processing_qty > 0
        elif current_filter == "warning":
            show_row = warning_qty > 0

        if not show_row:
            continue

        flower_rows.append({
            "product_code": stock.product_code,
            "product_name": stock.product_name,
            "unit": stock.unit or "",
            "flower_grade": stock.flower_grade or "",
            "supplier_name": stock.supplier_name or "Chưa có NCC",
            "latest_date": received_day,
            "oldest_days": days_in_stock,
            "total_qty": quantity,
            "new_qty": new_qty,
            "processing_qty": processing_qty,
            "warning_qty": warning_qty,
            "lot_code": stock.lot.lot_code if stock.lot else "Chưa có mã lô",
            "trace_url": f"/trace/{stock.lot.lot_code}/" if stock.lot else "#",
            "has_lot": True if stock.lot else False,
        })

    context = {
        "flower_rows": flower_rows,
        "total_all": total_all,
        "total_new": total_new,
        "total_processing": total_processing,
        "total_warning": total_warning,
        "current_filter": current_filter,
    }

    return render(request, "inventory/raw_inventory.html", context)

def material_inventory(request):
    q = request.GET.get('q', '')

    stocks = InventoryStock.objects.filter(
        warehouse_type='MATERIAL',
        product_group='VT'
    )

    if q:
        stocks = stocks.filter(
            Q(product_code__icontains=q) |
            Q(product_name__icontains=q)
        )

    stocks = grouped_stock_queryset(stocks)

    context = {
        'stocks': stocks,
        'q': q,
        'page_title': 'Kho vật tư',
        'page_desc': 'Danh sách vật tư đóng gói, thùng, đá gel và phụ liệu',
    }

    return render(request, 'inventory/stock_list.html', context)


def finished_inventory(request):
    from processing.models import ProcessingTicketItem

    q = request.GET.get('q', '')

    stocks = (
        InventoryStock.objects
        .filter(warehouse_type="FINISHED")
        .order_by("product_name")
    )

    if q:
        stocks = stocks.filter(
            Q(product_code__icontains=q) |
            Q(product_name__icontains=q)
        )

    grouped = {}

    for stock in stocks:

        key = (
            stock.product_code or "",
            stock.product_name or "",
        )

        if key not in grouped:

            grouped[key] = {
                "product_code": stock.product_code or "-",
                "product_name": stock.product_name or "-",
                "bunch_type": stock.flower_grade or "-",

                "total_finished_stems": 0,
                "total_finished_bunches": 0,

                "total_raw_stems": 0,

                "items": [],
                "latest_items": [],
            }

        grouped[key]["total_finished_stems"] += float(
            stock.quantity or 0
        )

    finished_groups = []
    total_all_finished_stems = 0

    for key, group in grouped.items():

        product_code = group["product_code"]

        processing_items = list(
            ProcessingTicketItem.objects
            .select_related("ticket")
            .filter(product_code=product_code)
            .order_by("-created_at")
        )

        group["items"] = processing_items
        group["latest_items"] = processing_items[:2]

        group["total_raw_stems"] = (
            InventoryStock.objects
            .filter(
                warehouse_type="RAW",
                product_code=product_code
            )
            .aggregate(total=Sum("quantity"))["total"] or 0
        )

        total_all_finished_stems += group["total_finished_stems"]

        finished_groups.append(group)

    finished_groups.sort(
        key=lambda x: x["product_name"]
    )

    context = {
        "finished_groups": finished_groups,
        "q": q,
        "total_all_finished_stems": total_all_finished_stems,
    }

    return render(
        request,
        "inventory/finished_inventory.html",
        context
    )


def sale_inventory(request):
    q = request.GET.get('q', '')

    stocks = InventoryStock.objects.filter(
        warehouse_type='SALE'
    )

    if q:
        stocks = stocks.filter(
            Q(product_code__icontains=q) |
            Q(product_name__icontains=q)
        )

    stocks = grouped_stock_queryset(stocks)

    context = {
        'stocks': stocks,
        'q': q,
        'page_title': 'Kho Sale',
        'page_desc': 'Sản phẩm sale đã lấy từ kho thành phẩm',
    }

    return render(request, 'inventory/stock_list.html', context)


def damaged_inventory(request):
    q = request.GET.get('q', '')

    stocks = InventoryStock.objects.filter(
        warehouse_type='DAMAGED'
    ).select_related("lot")

    if q:
        stocks = stocks.filter(
            Q(product_code__icontains=q) |
            Q(product_name__icontains=q)
        )

    for stock in stocks:
        stock.latest_stock_out = (
            InventoryStockOutItem.objects
            .select_related("stock_out")
            .filter(
                lot=stock.lot,
                destination_warehouse="DAMAGED"
            )
            .order_by("-created_at")
            .first()
        )

    context = {
        'stocks': stocks,
        'q': q,
        'page_title': 'Hoa/Lá phụ hủy',
        'page_desc': 'Hoa, lá phụ bị hủy hoặc không đạt tiêu chuẩn',
    }

    return render(request, 'inventory/stock_list.html', context)

def raw_material_inventory(request):
    today = timezone.localdate()

    items = (
    InventoryReceiptItem.objects
    .select_related("receipt", "lot")
    .filter(receipt__product_group__in=["HH", "HP"])
    .order_by("receipt__receipt_datetime")
)

    new_items = []
    need_process_items = []
    warning_items = []

    flower_totals = defaultdict(float)
    supplier_totals = defaultdict(float)

    total_new = 0
    total_need_process = 0
    total_warning = 0

    for item in items:
        receipt_date = timezone.localtime(item.receipt.receipt_datetime).date()
        days_in_stock = (today - receipt_date).days
        quantity = item.quantity or 0

        item.receipt_date_display = receipt_date
        item.days_in_stock = days_in_stock
        item.supplier_display = item.receipt.supplier_name or "Chưa có NCC"
        item.lot_code_display = item.lot.lot_code if item.lot else "Chưa có mã lô"
        item.trace_url = f"/trace/{item.lot.lot_code}/" if item.lot else "#"
        item.qr_url = item.lot.qr_code.url if item.lot and item.lot.qr_code else ""

        flower_totals[item.product_name] += quantity
        supplier_totals[item.supplier_display] += quantity

        if receipt_date == today:
            new_items.append(item)
            total_new += quantity
        else:
            need_process_items.append(item)
            total_need_process += quantity

        if days_in_stock >= 2:
            warning_items.append(item)
            total_warning += quantity

    context = {
        "new_items": new_items,
        "need_process_items": need_process_items,
        "warning_items": warning_items,

        "flower_totals": dict(flower_totals),
        "supplier_totals": dict(supplier_totals),

        "total_new": total_new,
        "total_need_process": total_need_process,
        "total_warning": total_warning,
        "total_all": total_new + total_need_process,
    }

    return render(request, "inventory/raw_material_inventory.html", context)

def api_stock_out_products(request):
    stock_out_type = request.GET.get("type", "")

    stocks = InventoryStock.objects.filter(quantity__gt=0).select_related("lot")

    if stock_out_type == "HH":
        stocks = stocks.filter(
            warehouse_type="FINISHED",
            product_group="HH"
        )

    elif stock_out_type == "HP":
        stocks = stocks.filter(
            warehouse_type="FINISHED",
            product_group="HP"
        )

    elif stock_out_type == "DIRECT_SALE":
        stocks = stocks.filter(
            warehouse_type="RAW",
            product_group__in=["HH", "HP"]
        )

    elif stock_out_type == "DAMAGED":
        stocks = stocks.filter(
            warehouse_type__in=["FINISHED", "SALE"],
            product_group__in=["HH", "HP"]
        )

    else:
        return JsonResponse({"products": []})

    data = []

    for stock in stocks.order_by("product_name", "received_date"):
        data.append({
            "stock_id": stock.id,
            "lot_id": stock.lot.id if stock.lot else "",
            "lot_code": stock.lot.lot_code if stock.lot else "Chưa có mã lô",

            "warehouse_type": stock.warehouse_type,
            "product_group": stock.product_group,

            "product_code": stock.product_code,
            "product_name": stock.product_name,
            "unit": stock.unit or "",
            "flower_grade": stock.flower_grade or "",

            "supplier_code": stock.supplier_code or "",
            "supplier_name": stock.supplier_name or "",

            "received_date": stock.received_date.strftime("%d/%m/%Y %H:%M") if stock.received_date else "",
            "quantity": stock.quantity or 0,
        })

    return JsonResponse({"products": data})

from django.views.decorators.csrf import csrf_exempt
@csrf_exempt
def create_stock_out(request):
    if request.method != "POST":
        return redirect("inventory:dashboard")

    stock_out_type = request.POST.get("stock_out_type")
    items_json = request.POST.get("stock_out_items_json")
    business_date = parse_business_date(request.POST.get("business_date"))

    if not stock_out_type:
        messages.error(request, "Vui lòng chọn loại xuất kho.")
        return redirect("inventory:dashboard")

    if not items_json:
        messages.error(request, "Vui lòng thêm ít nhất một sản phẩm xuất kho.")
        return redirect("inventory:dashboard")

    try:
        items = json.loads(items_json)
    except json.JSONDecodeError:
        messages.error(request, "Dữ liệu xuất kho không hợp lệ.")
        return redirect("inventory:dashboard")

    if not items:
        messages.error(request, "Vui lòng thêm ít nhất một sản phẩm xuất kho.")
        return redirect("inventory:dashboard")

    stock_out_code = f"XK{timezone.now().strftime('%Y%m%d%H%M%S')}"

    employee_code = ""
    employee_name = request.user.username if request.user.is_authenticated else ""
    employee_department = ""
    employee_position = ""

    if request.user.is_authenticated:
        user_employee = getattr(request.user, "employee_profile", None)

        if user_employee and user_employee.status == "approved":
            employee = user_employee.employee
            employee_code = employee.ma_nv or ""
            employee_name = employee.ho_ten or request.user.username
            employee_department = employee.bo_phan or ""
            employee_position = employee.chuc_vu or ""

    stock_out = InventoryStockOut.objects.create(
        stock_out_code=stock_out_code,
        stock_out_type=stock_out_type,
        employee_code=employee_code,
        employee_name=employee_name,
        employee_department=employee_department,
        employee_position=employee_position,
    )

    for item in items:
        stock_id = item.get("stock_id")
        quantity = float(item.get("quantity") or 0)

        if quantity <= 0:
            continue

        # Khóa dòng dữ liệu tồn kho nguồn để tránh xung đột khi nhiều máy cùng xuất một lúc
        source_stock = InventoryStock.objects.select_for_update().filter(id=stock_id).first()

        if not source_stock:
            messages.error(request, f"Không tìm thấy tồn kho nguồn cho {item.get('product_name')}.")
            return redirect("inventory:dashboard")

        if quantity > float(source_stock.quantity or 0):
            messages.error(
                request,
                f"{source_stock.product_name} chỉ còn {source_stock.quantity}, không thể xuất {quantity}."
            )
            return redirect("inventory:dashboard")

        if stock_out_type in ["HH", "HP", "DIRECT_SALE"]:
            destination_warehouse = "SALE"
            trace_action = "move_to_sale"
            trace_to_area = "Kho Sale"
            transaction_note = "Chuyển sang kho Sale"
        elif stock_out_type == "DAMAGED":
            destination_warehouse = "DAMAGED"
            trace_action = "destroy"
            trace_to_area = "Kho Hoa Hủy"
            transaction_note = "Xuất hoa hủy"
        else:
            destination_warehouse = "SALE"
            trace_action = "other"
            trace_to_area = "Kho khác"
            transaction_note = "Xuất kho"

        source_warehouse = source_stock.warehouse_type
        output_grade = item.get("output_grade") or source_stock.flower_grade or ""
        reason = item.get("reason") or ""

        stock_out_item = InventoryStockOutItem.objects.create(
            stock_out=stock_out,
            lot=source_stock.lot,
            source_stock_id=source_stock.id,
            source_warehouse=source_warehouse,
            destination_warehouse=destination_warehouse,
            product_group=source_stock.product_group,
            product_code=source_stock.product_code,
            product_name=source_stock.product_name,
            unit=source_stock.unit,
            original_grade=source_stock.flower_grade,
            output_grade=output_grade,
            supplier_code=source_stock.supplier_code,
            supplier_name=source_stock.supplier_name,
            received_date=source_stock.received_date,
            reason=reason,
            quantity=quantity,
        )

        source_stock.quantity -= quantity
        source_stock.save(update_fields=["quantity", "updated_at"])

        dest_stock, created = InventoryStock.objects.get_or_create(
            lot=source_stock.lot,
            warehouse_type=destination_warehouse,
            product_group=source_stock.product_group,
            product_code=source_stock.product_code,
            flower_grade=output_grade,
            supplier_code=source_stock.supplier_code,
            received_date=source_stock.received_date,
            defaults={
                "product_name": source_stock.product_name,
                "unit": source_stock.unit,
                "supplier_name": source_stock.supplier_name,
                "quantity": 0,
            }
        )

        dest_stock.product_name = source_stock.product_name
        dest_stock.unit = source_stock.unit
        dest_stock.supplier_name = source_stock.supplier_name
        dest_stock.quantity += quantity
        dest_stock.save()

        InventoryTransaction.objects.create(
            lot=source_stock.lot,
            transaction_code=f"GD{timezone.now().strftime('%Y%m%d%H%M%S%f')}",
            transaction_type="OUT",
            warehouse_type=source_warehouse,
            product_group=source_stock.product_group,
            product_code=source_stock.product_code,
            product_name=source_stock.product_name,
            unit=source_stock.unit,
            flower_grade=source_stock.flower_grade,
            supplier_code=source_stock.supplier_code,
            supplier_name=source_stock.supplier_name,
            received_date=source_stock.received_date,
            quantity=quantity,
            reference_code=stock_out.stock_out_code,
            business_date=business_date,
            note=transaction_note,
        )

        InventoryTransaction.objects.create(
            lot=source_stock.lot,
            transaction_code=f"GD{timezone.now().strftime('%Y%m%d%H%M%S%f')}",
            transaction_type="IN",
            warehouse_type=destination_warehouse,
            product_group=source_stock.product_group,
            product_code=source_stock.product_code,
            product_name=source_stock.product_name,
            unit=source_stock.unit,
            flower_grade=output_grade,
            supplier_code=source_stock.supplier_code,
            supplier_name=source_stock.supplier_name,
            received_date=source_stock.received_date,
            quantity=quantity,
            reference_code=stock_out.stock_out_code,
            business_date=business_date,
            note=transaction_note,
        )

        if source_stock.lot:
            from traceability.services import add_trace_log

            add_trace_log(
                lot=source_stock.lot,
                action=trace_action,
                quantity=quantity,
                from_area=source_warehouse,
                to_area=trace_to_area,
                employee_name=employee_name,
                related_code=stock_out.stock_out_code,
                business_date=business_date,
                note=f"{transaction_note}. Lý do: {reason}",
            )

    messages.success(request, f"Đã tạo phiếu xuất kho {stock_out_code} thành công.")
    return redirect("inventory:dashboard")

def api_pending_stock_in_orders(request):
    orders = (
        PurchaseOrder.objects
        .filter(status="waiting_stock_in")
        .prefetch_related(
            "items",
            "items__warehouse_check_draft"
        )
        .order_by("-created_at")
    )

    data = []

    for po in orders:
        supplier_names = set()

        for item in po.items.all():
            if item.supplier_name:
                supplier_names.add(item.supplier_name.strip())
            elif po.supplier_name:
                supplier_names.add(po.supplier_name.strip())

        supplier_display = (
            "Đơn hàng hỗn hợp"
            if len(supplier_names) >= 2
            else po.supplier_name
        )

        items = []

        for item in po.items.all():

            draft = getattr(
                item,
                "warehouse_check_draft",
                None
            )

            checked_quantity = (
                draft.checked_quantity
                if draft
                else (
                    item.warehouse_checked_quantity
                    or item.driver_received_quantity
                    or item.ordered_quantity
                    or 0
                )
            )

            checked_stems_per_bundle = (
                draft.stems_per_bundle
                if draft
                else (
                    item.stems_per_bundle
                    or 50
                )
            )

            items.append({
                "id": item.id,

                "product_name": item.product_name,
                "product_code": item.product_code,

                "supplier_name": (
                    item.supplier_name
                    or po.supplier_name
                    or ""
                ),

                "unit": item.unit or "cành",

                "stems_per_bundle": checked_stems_per_bundle,

                "ordered_quantity":
                    item.ordered_quantity or 0,

                "driver_received_quantity":
                    item.driver_received_quantity
                    or item.ordered_quantity
                    or 0,

                "warehouse_checked_quantity":
                    checked_quantity,

                "note":
                    item.note or "",
            })

        data.append({
            "id": po.id,

            "po_code": po.po_code,

            "supplier_display":
                supplier_display,

            "driver_name":
                po.driver_name
                or "Chưa có tài xế nhận",

            "vehicle_info":
                po.vehicle_info
                or "-",

            "status_display":
                po.get_status_display(),

            "items":
                items,
        })

    return JsonResponse({
        "success": True,
        "orders": data,
    })

@transaction.atomic
def create_stock_in_from_purchase_order(request, po_id):
    if request.method != "POST":
        return redirect("inventory:dashboard")

    po = (
        PurchaseOrder.objects
        .select_for_update()
        .prefetch_related("items")
        .filter(id=po_id, status="waiting_stock_in")
        .first()
    )

    if not po:
        messages.error(request, "Không tìm thấy đơn chờ nhập kho hoặc đơn đã được nhập.")
        return redirect("inventory:dashboard")

    stock_in_items = []

    for item in po.items.all():
        stock_in_qty = float(
            request.POST.get(f"stock_in_qty_{item.id}") or 0
        )

        stems_per_bundle = int(
            request.POST.get(f"stems_per_bundle_{item.id}") or item.stems_per_bundle or 50
        )

        if stock_in_qty <= 0:
            continue

        product = FlowerType.objects.filter(code=item.product_code).first()

        product_group = "HH"
        if product and product.category_type:
            product_group = product.category_type

        supplier_code = item.supplier_code or po.supplier_code or ""
        supplier_name = item.supplier_name or po.supplier_name or ""

        stock_in_items.append({
            "purchase_item": item,
            "product_group": product_group,
            "supplier_code": supplier_code,
            "supplier_name": supplier_name,
            "stock_in_qty": stock_in_qty,
            "stems_per_bundle": stems_per_bundle,
        })

    if not stock_in_items:
        messages.error(request, "Vui lòng nhập số lượng kho kiểm lớn hơn 0.")
        return redirect("inventory:dashboard")

    receipts_by_supplier = {}

    for data in stock_in_items:
        supplier_key = (
            data["supplier_code"],
            data["supplier_name"],
            data["product_group"],
        )

        if supplier_key not in receipts_by_supplier:
            receipt_code = f"NK{timezone.now().strftime('%Y%m%d%H%M%S%f')}"

            receipts_by_supplier[supplier_key] = InventoryReceipt.objects.create(
                receipt_code=receipt_code,
                product_group=data["product_group"],
                supplier_code=data["supplier_code"],
                supplier_name=data["supplier_name"],
                note=f"Nhập kho từ đơn vận chuyển {po.po_code}",
            )

        receipt = receipts_by_supplier[supplier_key]
        purchase_item = data["purchase_item"]

        receipt_item = InventoryReceiptItem.objects.create(
            receipt=receipt,
            product_code=purchase_item.product_code,
            product_name=purchase_item.product_name,
            unit=purchase_item.unit or "cành",
            flower_grade="",
            quantity=data["stock_in_qty"],
        )

        lot = create_trace_lot_from_receipt_item(
            receipt,
            receipt_item,
            request=request
        )

        create_trace_bundles_from_purchase_item(
            lot=lot,
            product_code=purchase_item.product_code,
            product_name=purchase_item.product_name,
            supplier_code=data["supplier_code"],
            supplier_name=data["supplier_name"],
            quantity=data["stock_in_qty"],
            stems_per_bundle=data["stems_per_bundle"],
            request=request,
        )

        warehouse_type = "MATERIAL" if data["product_group"] == "VT" else "RAW"

        stock, created = InventoryStock.objects.get_or_create(
            lot=lot,
            warehouse_type=warehouse_type,
            product_group=data["product_group"],
            product_code=purchase_item.product_code,
            flower_grade="",
            supplier_code=data["supplier_code"],
            received_date=receipt.receipt_datetime,
            defaults={
                "product_name": purchase_item.product_name,
                "unit": purchase_item.unit or "cành",
                "supplier_name": data["supplier_name"],
                "quantity": 0,
            }
        )

        stock.product_name = purchase_item.product_name
        stock.unit = purchase_item.unit or "cành"
        stock.supplier_name = data["supplier_name"]
        stock.quantity += data["stock_in_qty"]
        stock.save()

        InventoryTransaction.objects.create(
            lot=lot,
            transaction_code=f"GD{timezone.now().strftime('%Y%m%d%H%M%S%f')}",
            transaction_type="IN",
            warehouse_type=warehouse_type,
            product_group=data["product_group"],
            product_code=purchase_item.product_code,
            product_name=purchase_item.product_name,
            unit=purchase_item.unit or "cành",
            flower_grade="",
            supplier_code=data["supplier_code"],
            supplier_name=data["supplier_name"],
            received_date=receipt.receipt_datetime,
            quantity=data["stock_in_qty"],
            reference_code=receipt.receipt_code,
            note=f"Nhập kho từ đơn vận chuyển {po.po_code}",
        )

        purchase_item.received_quantity = data["stock_in_qty"]
        purchase_item.stems_per_bundle = data["stems_per_bundle"]
        purchase_item.save()

    po.status = "received"
    po.received_date = timezone.localdate()

    if request.user.is_authenticated:
        po.warehouse_receiver = request.user.get_full_name() or request.user.username
        po.warehouse_receiver_code = request.user.username

    po.save()

    messages.success(
        request,
        f"Đã nhập kho đơn {po.po_code}, tạo mã lô và QR cho từng bó thành công."
    )

    return redirect("inventory:dashboard")

@transaction.atomic
def create_bundle_labels_from_purchase_order(request, po_id):
    if request.method != "POST":
        return redirect("inventory:dashboard")

    po = (
        PurchaseOrder.objects
        .select_for_update()
        .prefetch_related("items")
        .filter(id=po_id, status="waiting_stock_in")
        .first()
    )

    if not po:
        messages.error(request, "Không tìm thấy đơn chờ nhập kho hoặc đơn đã xử lý.")
        return redirect("inventory:dashboard")

    created_receipts = []

    for item in po.items.all():
        stock_in_qty = float(
            item.warehouse_checked_quantity
            or request.POST.get(f"stock_in_qty_{item.id}")
            or item.driver_received_quantity
            or item.ordered_quantity
            or 0
        )

        stems_per_bundle = int(
            request.POST.get(f"stems_per_bundle_{item.id}") or item.stems_per_bundle or 50
        )

        if stock_in_qty <= 0:
            continue

        product = FlowerType.objects.filter(code=item.product_code).first()
        product_group = product.category_type if product and product.category_type else "HH"

        supplier_code = item.supplier_code or po.supplier_code or ""
        supplier_name = item.supplier_name or po.supplier_name or ""

        receipt_code = f"NK-TEM{timezone.now().strftime('%Y%m%d%H%M%S%f')}"

        receipt = InventoryReceipt.objects.create(
            receipt_code=receipt_code,
            product_group=product_group,
            supplier_code=supplier_code,
            supplier_name=supplier_name,
            source_po_code=po.po_code,
            status="DRAFT_LABEL_CREATED",
            note=f"Tạo QR bó trước nhập kho từ đơn {po.po_code}",
        )

        receipt_item = InventoryReceiptItem.objects.create(
            receipt=receipt,
            product_code=item.product_code,
            product_name=item.product_name,
            unit=item.unit or "cành",
            flower_grade="",
            quantity=stock_in_qty,
        )

        lot = create_trace_lot_from_receipt_item(
            receipt,
            receipt_item,
            request=request
        )

        create_trace_bundles_from_purchase_item(
            lot=lot,
            product_code=item.product_code,
            product_name=item.product_name,
            supplier_code=supplier_code,
            supplier_name=supplier_name,
            quantity=stock_in_qty,
            stems_per_bundle=stems_per_bundle,
            request=request,
        )

        item.received_quantity = stock_in_qty
        item.stems_per_bundle = stems_per_bundle
        item.save()

        created_receipts.append(receipt.receipt_code)

    if not created_receipts:
        messages.error(request, "Vui lòng nhập số lượng kho kiểm lớn hơn 0.")
        return redirect("inventory:dashboard")

    messages.success(
        request,
        f"Đã tạo QR bó cho đơn {po.po_code}. Nhân viên kho có thể in tem và dán lên bó."
    )

    return redirect("inventory:dashboard")

@require_POST
def api_save_warehouse_checked_quantity(request):
    try:
        data = json.loads(request.body.decode("utf-8"))

        item_id = data.get("item_id")
        quantity = float(data.get("quantity") or 0)
        stems_per_bundle = int(data.get("stems_per_bundle") or 50)

        item = PurchaseOrderItem.objects.get(id=item_id)

        draft, created = PurchaseWarehouseCheckDraft.objects.get_or_create(
            purchase_order_item=item,
            defaults={
                "checked_quantity": quantity,
                "stems_per_bundle": stems_per_bundle,
            }
        )

        draft.checked_quantity = quantity
        draft.stems_per_bundle = stems_per_bundle
        draft.save()

        return JsonResponse({"success": True})

    except Exception as e:
        return JsonResponse({
            "success": False,
            "message": str(e)
        }, status=400)
    
@transaction.atomic
def quick_inventory_check(request):
    if request.method != "POST":
        return redirect("inventory:dashboard")

    default_product_group = request.POST.get("check_product_group")
    default_warehouse_type = request.POST.get("check_warehouse_type")
    items_json = request.POST.get("check_items_json")
    business_date = parse_business_date(request.POST.get("business_date"))

    if not items_json:
        messages.error(request, "Vui lòng thêm ít nhất một sản phẩm kiểm kê.")
        return redirect("inventory:dashboard")

    try:
        items = json.loads(items_json)
    except json.JSONDecodeError:
        messages.error(request, "Dữ liệu kiểm kê không hợp lệ.")
        return redirect("inventory:dashboard")

    if not items:
        messages.error(request, "Vui lòng thêm ít nhất một sản phẩm kiểm kê.")
        return redirect("inventory:dashboard")

    check_code = f"KK{timezone.now().strftime('%Y%m%d%H%M%S')}"

    def write_check_transaction(stock, quantity, note):
        InventoryTransaction.objects.create(
            lot=stock.lot,
            transaction_code=f"KK{timezone.now().strftime('%Y%m%d%H%M%S%f')}",
            transaction_type="CHECK",
            warehouse_type=stock.warehouse_type,
            product_group=stock.product_group,
            product_code=stock.product_code or "-",
            product_name=stock.product_name or "-",
            unit=stock.unit or "-",
            flower_grade=normalize_flower_grade(stock.flower_grade),
            supplier_code=stock.supplier_code or "",
            supplier_name=stock.supplier_name or "-",
            received_date=stock.received_date,
            quantity=quantity,
            reference_code=check_code,
            business_date=business_date,
            note=note,
        )

    for item in items:
        product_group = item.get("product_group") or default_product_group
        warehouse_type = item.get("warehouse_type") or default_warehouse_type
        sale_type = str(item.get("sale_type") or "").strip()

        product_code = str(item.get("product_code") or "").strip()
        product_name = str(item.get("product_name") or "").strip()
        unit = str(item.get("unit") or "").strip() or "-"

        final_quantity = float(item.get("quantity") or 0)

        if not product_group or not warehouse_type or not product_name:
            continue

        if final_quantity < 0:
            continue

        if warehouse_type in ["SALE", "DAMAGED"] and sale_type not in ["50c", "100c"]:
            messages.error(
                request,
                "Kho Sale hoặc Hoa/Lá phụ hủy phải chọn 50c hoặc 100c."
            )
            return redirect("inventory:dashboard")

        product = FlowerType.objects.filter(code=product_code).first()

        if product:
            product_name = product.name or product_name
            unit = product.unit or unit or "-"

        flower_grade = normalize_flower_grade(
            sale_type if warehouse_type in ["SALE", "DAMAGED"] else ""
        )

        target_stock = get_or_create_normalized_stock(
            warehouse_type=warehouse_type,
            product_group=product_group,
            product_code=product_code,
            product_name=product_name,
            unit=unit,
            flower_grade=flower_grade,
        )

        old_target_quantity = float(target_stock.quantity or 0)
        diff_quantity = final_quantity - old_target_quantity

        target_stock.quantity = final_quantity
        target_stock.save()

        write_check_transaction(
            target_stock,
            diff_quantity,
            (
                f"Kiểm kê chốt tồn cuối. "
                f"Tồn hệ thống: {old_target_quantity:g}. "
                f"Tồn thực tế: {final_quantity:g}. "
                f"Chênh lệch: {diff_quantity:g}"
            ),
        )

        source_warehouse = None
        source_flower_grade = ""

        if warehouse_type == "SALE":
            if sale_type == "50c":
                source_warehouse = "FINISHED"
            elif sale_type == "100c":
                source_warehouse = "RAW"

        elif warehouse_type == "DAMAGED":
            source_warehouse = "SALE"
            source_flower_grade = normalize_flower_grade(sale_type)

        if source_warehouse and diff_quantity != 0:
            source_stock = get_or_create_normalized_stock(
                warehouse_type=source_warehouse,
                product_group=product_group,
                product_code=product_code,
                product_name=product_name,
                unit=unit,
                flower_grade=source_flower_grade,
            )

            old_source_quantity = float(source_stock.quantity or 0)
            new_source_quantity = old_source_quantity - diff_quantity

            source_stock.quantity = new_source_quantity
            source_stock.save()

            write_check_transaction(
                source_stock,
                -diff_quantity,
                (
                    f"Kiểm kê tự động điều chỉnh nguồn. "
                    f"Kho đích: {warehouse_type} {sale_type}. "
                    f"Tồn nguồn trước: {old_source_quantity:g}. "
                    f"Điều chỉnh: {-diff_quantity:g}. "
                    f"Tồn nguồn mới: {new_source_quantity:g}"
                ),
            )

    messages.success(
        request,
        f"Đã hoàn thành kiểm kê {check_code}. Tồn kho đã được đồng bộ theo số lượng thực tế."
    )

    return redirect("inventory:dashboard")


def inventory_stock_report(request):
    selected_date_str = request.GET.get("date", "")

    if selected_date_str:
        try:
            selected_date = timezone.datetime.strptime(
                selected_date_str,
                "%Y-%m-%d"
            ).date()
        except ValueError:
            selected_date = timezone.localdate()
    else:
        selected_date = timezone.localdate()

    q = request.GET.get("q", "").strip()
    warehouse_filter = request.GET.get("warehouse", "")
    group_filter = request.GET.get("group", "")

    stocks = (
        InventoryStock.objects
        .all()
        .order_by("product_name", "warehouse_type", "flower_grade")
    )

    if q:
        stocks = stocks.filter(
            Q(product_code__icontains=q) |
            Q(product_name__icontains=q)
        )

    if warehouse_filter:
        stocks = stocks.filter(warehouse_type=warehouse_filter)

    if group_filter:
        stocks = stocks.filter(product_group=group_filter)

    report_map = {}

    for stock in stocks:
        key = (
            stock.product_code or "",
            stock.product_name or "",
            stock.product_group or "",
            stock.unit or "",
        )

        if key not in report_map:
            report_map[key] = {
                "product_code": stock.product_code or "-",
                "product_name": stock.product_name or "-",
                "product_group": stock.get_product_group_display(),
                "unit": stock.unit or "-",

                "raw_qty": 0,
                "finished_qty": 0,
                "sale_50_qty": 0,
                "sale_100_qty": 0,
                "sale_other_qty": 0,
                "damaged_50_qty": 0,
                "damaged_100_qty": 0,
                "damaged_other_qty": 0,
                "material_qty": 0,
                "total_qty": 0,
            }

        qty = float(stock.quantity or 0)

        if stock.warehouse_type == "RAW":
            report_map[key]["raw_qty"] += qty

        elif stock.warehouse_type == "FINISHED":
            report_map[key]["finished_qty"] += qty

        elif stock.warehouse_type == "SALE":
            if stock.flower_grade == "50c":
                report_map[key]["sale_50_qty"] += qty
            elif stock.flower_grade == "100c":
                report_map[key]["sale_100_qty"] += qty
            else:
                report_map[key]["sale_other_qty"] += qty

        elif stock.warehouse_type == "DAMAGED":
            if stock.flower_grade == "50c":
                report_map[key]["damaged_50_qty"] += qty
            elif stock.flower_grade == "100c":
                report_map[key]["damaged_100_qty"] += qty
            else:
                report_map[key]["damaged_other_qty"] += qty

        elif stock.warehouse_type == "MATERIAL":
            report_map[key]["material_qty"] += qty

    for row in report_map.values():
        row["total_qty"] = (
            row["raw_qty"]
            + row["finished_qty"]
            + row["sale_50_qty"]
            + row["sale_100_qty"]
            + row["sale_other_qty"]
            + row["material_qty"]
        )

    report_rows = sorted(
        report_map.values(),
        key=lambda x: x["product_name"]
    )

    total_stock_qty = sum(row["total_qty"] for row in report_rows)
    total_product_count = len(report_rows)

    day_start = timezone.make_aware(
        timezone.datetime.combine(
            selected_date,
            timezone.datetime.min.time()
        )
    )

    day_end = timezone.make_aware(
        timezone.datetime.combine(
            selected_date,
            timezone.datetime.max.time()
        )
    )

    today_in_transactions = InventoryTransaction.objects.filter(
        transaction_type="IN",
        created_at__range=(day_start, day_end)
    )

    today_out_transactions = InventoryTransaction.objects.filter(
        transaction_type="OUT",
        created_at__range=(day_start, day_end)
    )

    today_check_transactions = InventoryTransaction.objects.filter(
        transaction_type="CHECK",
        created_at__range=(day_start, day_end)
    )

    today_created_orders = PurchaseOrder.objects.filter(
        created_at__range=(day_start, day_end)
    ).count()

    today_exported_orders = InventoryStockOut.objects.filter(
        created_at__range=(day_start, day_end)
    ).count()

    total_in_today = (
        today_in_transactions.aggregate(total=Sum("quantity"))["total"] or 0
    )

    total_out_today = (
        today_out_transactions.aggregate(total=Sum("quantity"))["total"] or 0
    )

    total_check_today = (
        today_check_transactions.aggregate(total=Sum("quantity"))["total"] or 0
    )

    top_in_products = (
        today_in_transactions
        .values("product_code", "product_name", "unit")
        .annotate(total_quantity=Sum("quantity"))
        .order_by("-total_quantity")[:10]
    )

    top_out_products = (
        today_out_transactions
        .values("product_code", "product_name", "unit")
        .annotate(total_quantity=Sum("quantity"))
        .order_by("-total_quantity")[:10]
    )

    warehouse_totals = (
        stocks
        .values("warehouse_type")
        .annotate(total_quantity=Sum("quantity"))
        .order_by("warehouse_type")
    )

    context = {
        "q": q,
        "warehouse_filter": warehouse_filter,
        "group_filter": group_filter,

        "report_rows": report_rows,

        "total_stock_qty": total_stock_qty,
        "total_product_count": total_product_count,

        "today_created_orders": today_created_orders,
        "today_exported_orders": today_exported_orders,
        "total_in_today": total_in_today,
        "total_out_today": total_out_today,
        "total_check_today": total_check_today,

        "top_in_products": top_in_products,
        "top_out_products": top_out_products,
        "warehouse_totals": warehouse_totals,

        "warehouse_choices": InventoryStock.WAREHOUSE_CHOICES,
        "product_group_choices": InventoryStock.PRODUCT_GROUP_CHOICES,
        "selected_date": selected_date,
    }

    return render(
        request,
        "inventory/inventory_stock_report.html",
        context
    )

def export_inventory_stock_report_excel(request):
    import os
    from django.conf import settings
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    selected_date_str = request.GET.get("date", "")

    if selected_date_str:
        try:
            selected_date = timezone.datetime.strptime(selected_date_str, "%Y-%m-%d").date()
        except ValueError:
            selected_date = timezone.localdate()
    else:
        selected_date = timezone.localdate()

    q = request.GET.get("q", "").strip()
    warehouse_filter = request.GET.get("warehouse", "")
    group_filter = request.GET.get("group", "")

    stocks = InventoryStock.objects.all().order_by(
        "product_name",
        "warehouse_type",
        "flower_grade"
    )

    if q:
        stocks = stocks.filter(
            Q(product_code__icontains=q) |
            Q(product_name__icontains=q)
        )

    if warehouse_filter:
        stocks = stocks.filter(warehouse_type=warehouse_filter)

    if group_filter:
        stocks = stocks.filter(product_group=group_filter)

    report_map = {}

    for stock in stocks:
        key = (
            stock.product_code or "",
            stock.product_name or "",
            stock.product_group or "",
            stock.unit or "",
        )

        if key not in report_map:
            report_map[key] = {
                "product_code": stock.product_code or "-",
                "product_name": stock.product_name or "-",
                "product_group": stock.get_product_group_display(),
                "unit": stock.unit or "-",
                "raw_qty": 0,
                "finished_qty": 0,
                "sale_50_qty": 0,
                "sale_100_qty": 0,
                "damaged_50_qty": 0,
                "damaged_100_qty": 0,
                "material_qty": 0,
                "total_qty": 0,
            }

        qty = stock.quantity or 0

        if stock.warehouse_type == "RAW":
            report_map[key]["raw_qty"] += qty

        elif stock.warehouse_type == "FINISHED":
            report_map[key]["finished_qty"] += qty

        elif stock.warehouse_type == "SALE":
            if stock.flower_grade == "50c":
                report_map[key]["sale_50_qty"] += qty
            elif stock.flower_grade == "100c":
                report_map[key]["sale_100_qty"] += qty

        elif stock.warehouse_type == "DAMAGED":
            if stock.flower_grade == "50c":
                report_map[key]["damaged_50_qty"] += qty
            elif stock.flower_grade == "100c":
                report_map[key]["damaged_100_qty"] += qty

        elif stock.warehouse_type == "MATERIAL":
            report_map[key]["material_qty"] += qty

    for row in report_map.values():
        row["total_qty"] = (
            row["raw_qty"]
            + row["finished_qty"]
            + row["sale_50_qty"]
            + row["sale_100_qty"]
            + row["material_qty"]
        )

    report_rows = sorted(report_map.values(), key=lambda x: x["product_name"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Bao cao ton kho"

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A12"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    green_fill = PatternFill("solid", fgColor="15803D")
    dark_green_fill = PatternFill("solid", fgColor="14532D")
    light_green_fill = PatternFill("solid", fgColor="DCFCE7")
    pale_green_fill = PatternFill("solid", fgColor="ECFDF5")
    yellow_fill = PatternFill("solid", fgColor="FEF3C7")
    blue_fill = PatternFill("solid", fgColor="DBEAFE")
    orange_fill = PatternFill("solid", fgColor="FFEDD5")
    gray_fill = PatternFill("solid", fgColor="F8FAFC")

    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=18, bold=True, color="14532D")
    subtitle_font = Font(size=12, bold=True, color="166534")
    report_title_font = Font(size=16, bold=True, color="0F172A")
    bold_font = Font(bold=True)
    small_gray_font = Font(size=10, color="64748B")

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    logo_path = os.path.join(settings.BASE_DIR, "static", "images", "logon.png")
    if os.path.exists(logo_path):
        logo = ExcelImage(logo_path)
        logo.width = 85
        logo.height = 85
        ws.add_image(logo, "A1")

    ws.merge_cells("B1:M1")
    ws["B1"] = "CÔNG TY TNHH QUỲNH PHƯƠNG ĐÀ LẠT"
    ws["B1"].font = title_font
    ws["B1"].alignment = center

    ws.merge_cells("B2:M2")
    ws["B2"] = "QUỲNH PHƯƠNG FLOWER EXPORT SYSTEM"
    ws["B2"].font = subtitle_font
    ws["B2"].alignment = center

    ws.merge_cells("B3:M3")
    ws["B3"] = "BÁO CÁO TỒN KHO CHI TIẾT THEO SẢN PHẨM"
    ws["B3"].font = report_title_font
    ws["B3"].alignment = center

    ws.merge_cells("B4:M4")
    ws["B4"] = f"Ngày báo cáo: {selected_date.strftime('%d/%m/%Y')}"
    ws["B4"].font = small_gray_font
    ws["B4"].alignment = center

    exporter = "-"
    if request.user.is_authenticated:
        exporter = request.user.get_full_name() or request.user.username

    summary_data = [
        ("Tổng sản phẩm", len(report_rows)),
        ("Tổng tồn", sum(row["total_qty"] for row in report_rows)),
        ("Người xuất", exporter),
        ("Thời gian xuất", timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")),
    ]

    summary_start = 6

    for i, (label, value) in enumerate(summary_data):
        col = 1 + i * 3

        ws.merge_cells(start_row=summary_start, start_column=col, end_row=summary_start, end_column=col + 1)
        label_cell = ws.cell(row=summary_start, column=col, value=label)
        label_cell.fill = dark_green_fill
        label_cell.font = white_font
        label_cell.alignment = center
        label_cell.border = border

        ws.merge_cells(start_row=summary_start + 1, start_column=col, end_row=summary_start + 1, end_column=col + 1)
        value_cell = ws.cell(row=summary_start + 1, column=col, value=value)
        value_cell.fill = pale_green_fill
        value_cell.font = Font(size=13, bold=True, color="14532D")
        value_cell.alignment = center
        value_cell.border = border

    ws.merge_cells("A9:M9")
    ws["A9"] = "BẢNG TỒN KHO CHI TIẾT"
    ws["A9"].font = Font(size=13, bold=True, color="14532D")
    ws["A9"].alignment = left

    headers = [
        "STT",
        "Mã SP",
        "Tên sản phẩm",
        "Nhóm hàng",
        "ĐVT",
        "Kho nguyên liệu",
        "Kho thành phẩm",
        "Sale 50c",
        "Sale 100c",
        "Hủy 50c",
        "Hủy 100c",
        "Kho vật tư",
        "Tổng tồn",
    ]

    start_row = 11

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = wrap_center
        cell.border = border

    for index, row in enumerate(report_rows, 1):
        r = start_row + index

        values = [
            index,
            row["product_code"],
            row["product_name"],
            row["product_group"],
            row["unit"],
            row["raw_qty"],
            row["finished_qty"],
            row["sale_50_qty"],
            row["sale_100_qty"],
            row["damaged_50_qty"],
            row["damaged_100_qty"],
            row["material_qty"],
            row["total_qty"],
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.border = border
            cell.alignment = right if col >= 6 else left

            if col in [1, 2, 3, 4, 5]:
                cell.fill = gray_fill if index % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

            if col in [8, 9]:
                cell.fill = blue_fill

            elif col in [10, 11]:
                cell.fill = orange_fill

            elif col == 13:
                cell.fill = light_green_fill
                cell.font = bold_font

            if col == 3:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    total_row = start_row + len(report_rows) + 1

    ws.cell(row=total_row, column=1, value="TỔNG CỘNG")
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)

    for col in range(1, 14):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = yellow_fill
        cell.font = bold_font
        cell.border = border
        cell.alignment = center if col <= 5 else right

    for col in [6, 7, 8, 9, 10, 11, 12, 13]:
        col_letter = get_column_letter(col)
        ws.cell(
            row=total_row,
            column=col,
            value=f"=SUM({col_letter}{start_row + 1}:{col_letter}{total_row - 1})"
        )

    signature_row = total_row + 3

    ws.merge_cells(start_row=signature_row, start_column=2, end_row=signature_row, end_column=4)
    ws.cell(signature_row, 2, "Người lập báo cáo").font = bold_font
    ws.cell(signature_row, 2).alignment = center

    ws.merge_cells(start_row=signature_row, start_column=6, end_row=signature_row, end_column=8)
    ws.cell(signature_row, 6, "Thủ kho").font = bold_font
    ws.cell(signature_row, 6).alignment = center

    ws.merge_cells(start_row=signature_row, start_column=10, end_row=signature_row, end_column=12)
    ws.cell(signature_row, 10, "Ban giám đốc").font = bold_font
    ws.cell(signature_row, 10).alignment = center

    widths = {
        "A": 8,
        "B": 16,
        "C": 38,
        "D": 18,
        "E": 16,
        "F": 16,
        "G": 16,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 14,
        "L": 14,
        "M": 14,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row_num in range(1, total_row + 8):
        ws.row_dimensions[row_num].height = 24

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 26
    ws.row_dimensions[start_row].height = 38

    ws.auto_filter.ref = f"A{start_row}:M{total_row}"
    ws.print_title_rows = f"{start_row}:{start_row}"
    ws.oddFooter.center.text = "QUỲNH PHƯƠNG FLOWER EXPORT SYSTEM"
    ws.oddFooter.right.text = "Trang &[Page]/&[Pages]"

    file_name = f"bao_cao_ton_kho_{selected_date.strftime('%Y%m%d')}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'

    wb.save(response)
    return response