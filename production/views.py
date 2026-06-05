from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q, Sum
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse

import production
from .services import sync_production_items_from_order
from django.http import JsonResponse
from traceability.services import add_trace_log

import os
from django.conf import settings
from openpyxl import load_workbook

import qrcode
import base64
from io import BytesIO

from django.template.loader import get_template
from xhtml2pdf import pisa

from orders.models import Order
from categories.models import FlowerType
from .models import (
    ProductionOrder,
    ProductionItem,
    PackingBox,
    PackingBoxItem,
    PackingStemRule,
    BoxCapacity,
    PackingIndex,
)
from .services import create_production_from_order



def get_default_business_date():
    now = timezone.localtime(timezone.now())

    # Từ 00:00 đến trước 06:00 vẫn tính cho ngày hôm trước
    if now.hour < 6:
        return now.date() - timezone.timedelta(days=1)

    return now.date()

def get_flower_by_item(item):
    if item.product_code:
        flower = FlowerType.objects.filter(code__iexact=item.product_code).first()
        if flower:
            return flower

    if item.product_name:
        return FlowerType.objects.filter(name__icontains=item.product_name).first()

    return None


def get_packing_index_by_item(item):
    if item.product_code:
        packing_index = PackingIndex.objects.filter(
            flower__code__iexact=item.product_code,
            is_active=True
        ).first()

        if packing_index:
            return packing_index

    if item.product_name:
        packing_index = PackingIndex.objects.filter(
            flower__name__icontains=item.product_name,
            is_active=True
        ).first()

        if packing_index:
            return packing_index

    return None


def recalculate_box(box):
    total_bunches = 0
    total_stems = 0
    total_nw = 0
    total_index = 0

    for item in box.items.all():
        total_bunches += item.bunches or 0
        total_stems += item.stems or 0
        total_nw += item.nw or 0
        total_index += item.total_index or 0

    box.total_bunches = total_bunches
    box.total_stems = total_stems
    box.nw = total_nw
    box.total_index = total_index

    if box.box_type:
        box.capacity_index = box.box_type.capacity_index
        box.gw = total_nw + box.box_type.box_weight

        diff = box.capacity_index - total_index

        if diff < 0:
            box.status = "over"
        elif diff <= 2:
            box.status = "full"
        else:
            box.status = "lack"
    else:
        box.gw = total_nw
        box.status = "draft"

    box.save()


def production_list(request):
    q = request.GET.get("q", "")
    status = request.GET.get("status", "all")

    received_order_ids = ProductionOrder.objects.values_list("order_id", flat=True)

    pending_orders = (
        Order.objects
        .prefetch_related("items")
        .exclude(id__in=received_order_ids)
        .order_by("-order_time")
    )

    production_orders = (
        ProductionOrder.objects
        .prefetch_related("items", "boxes")
        .all()
        .order_by("-created_at")
    )

    if q:
        pending_orders = pending_orders.filter(
            Q(order_code__icontains=q) |
            Q(customer_name__icontains=q) |
            Q(items__product_name__icontains=q)
        ).distinct()

        production_orders = production_orders.filter(
            Q(order_code__icontains=q) |
            Q(customer_name__icontains=q) |
            Q(items__product_name__icontains=q)
        ).distinct()

    if status != "all":
        production_orders = production_orders.filter(status=status)

    context = {
        "pending_orders": pending_orders,
        "production_orders": production_orders,
        "q": q,
        "status": status,
        "total_orders": ProductionOrder.objects.count(),
        "pending_count": pending_orders.count(),
        "waiting_orders": ProductionOrder.objects.filter(status="waiting").count(),
        "processing_orders": ProductionOrder.objects.filter(status="processing").count(),
        "packing_orders": ProductionOrder.objects.filter(status="packing").count(),
        "completed_orders": ProductionOrder.objects.filter(status="completed").count(),
    }

    return render(request, "production/production_list.html", context)


def receive_order(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    create_production_from_order(order)

    messages.success(request, f"Đã nhận đơn {order.order_code} vào sản xuất.")
    return redirect("production:production_list")


def update_production_status(request, production_id, new_status):
    production = get_object_or_404(ProductionOrder, id=production_id)

    allowed_status = ["waiting", "processing", "packing", "completed", "cancelled"]

    if new_status not in allowed_status:
        messages.error(request, "Trạng thái sản xuất không hợp lệ.")
        return redirect("production:production_list")

    production.status = new_status

    if new_status == "processing" and not production.started_at:
        production.started_at = timezone.now()

    if new_status == "completed":
        production.completed_at = timezone.now()

    production.save()

    messages.success(request, f"Đã cập nhật trạng thái {production.production_code}.")
    return redirect("production:production_list")


def packing_detail(request, production_id):
    production = get_object_or_404(
        ProductionOrder.objects.prefetch_related(
            "items",
            "boxes__items",
            "items__box_items",
        ),
        id=production_id
    )

    sync_production_items_from_order(production)

    stem_rules = PackingStemRule.objects.all().order_by("-stems_quantity")
    box_types = BoxCapacity.objects.filter(is_active=True).order_by("-capacity_index")

    packing_rows = []

    for item in production.items.all():
        packed_qty = item.box_items.aggregate(total=Sum("stems"))["total"] or 0
        remaining_qty = max(float(item.required_quantity or 0) - float(packed_qty or 0), 0)

        item.packed_quantity = packed_qty
        item.remaining_quantity = remaining_qty

        if remaining_qty > 0:
            packing_rows.append({
                "item": item,
                "packed_qty": packed_qty,
                "remaining_qty": remaining_qty,
            })

    unpacked_items = [row["item"] for row in packing_rows]

    return render(request, "production/packing_detail.html", {
        "production": production,
        "stem_rules": stem_rules,
        "box_types": box_types,
        "unpacked_items": unpacked_items,
        "packing_rows": packing_rows,
        "has_unpacked_items": len(packing_rows) > 0,
    })


def update_packing_quantities(request, production_id):
    production = get_object_or_404(ProductionOrder, id=production_id)

    if request.method != "POST":
        return redirect("production:packing_detail", production_id=production.id)
    print("===== CREATE BOX POST =====")
    print(request.POST)

    for item in production.items.all():
        new_qty = request.POST.get(f"qty_{item.id}")

        if new_qty is None:
            continue

        new_qty = float(new_qty or 0)

        original_total = float(item.standard_quantity or 0) + float(item.sale_quantity or 0)

        packed_qty = item.box_items.aggregate(total=Sum("stems"))["total"] or 0
        packed_qty = float(packed_qty or 0)

        if new_qty > original_total:
            messages.error(
                request,
                f"{item.product_name}: SL thực tế không được vượt quá tổng đơn ({original_total} cành)."
            )
            return redirect("production:packing_detail", production_id=production.id)

        if new_qty < packed_qty:
            messages.error(
                request,
                f"{item.product_name}: SL thực tế không được nhỏ hơn số đã đóng ({packed_qty} cành)."
            )
            return redirect("production:packing_detail", production_id=production.id)

        item.required_quantity = new_qty
        item.save(update_fields=["required_quantity"])

    messages.success(request, "Đã cập nhật số lượng thực tế để đóng gói.")
    return redirect("production:packing_detail", production_id=production.id)

def get_available_stock(warehouse_type, product_code, product_name=""):
    from inventory.models import InventoryStock

    stocks = InventoryStock.objects.filter(
        warehouse_type=warehouse_type,
        quantity__gt=0,
    )

    if product_code:
        stocks = stocks.filter(product_code=product_code)
    elif product_name:
        stocks = stocks.filter(product_name__icontains=product_name)

    return stocks.order_by("received_date", "id").first()

def create_packing_box(request, production_id):
    from inventory.models import InventoryStock, InventoryTransaction

    production = get_object_or_404(ProductionOrder, id=production_id)

    if request.method != "POST":
        return redirect("production:packing_detail", production_id=production.id)

    box_type_id = request.POST.get("box_type_id")
    stem_rule_id = request.POST.get("stem_rule_id")

    box_type = get_object_or_404(BoxCapacity, id=box_type_id)
    stem_rule = get_object_or_404(PackingStemRule, id=stem_rule_id)

    stems_per_bunch = float(stem_rule.stems_quantity or 1)
    items_to_pack = []

    def get_business_date_for_packing():
        now = timezone.localtime(timezone.now())

        if now.hour < 6:
            return now.date() - timezone.timedelta(days=1)

        return now.date()

    def create_packing_out_transaction(stock, quantity, box, production, note):
        if stock is None:
            return

        quantity = float(quantity or 0)

        if quantity <= 0:
            return

        InventoryTransaction.objects.create(
            lot=stock.lot,
            transaction_code=f"DG{timezone.now().strftime('%Y%m%d%H%M%S%f')}",
            transaction_type="OUT",
            warehouse_type=stock.warehouse_type,
            product_group=stock.product_group,
            product_code=stock.product_code or "-",
            product_name=stock.product_name or "-",
            unit=stock.unit or "-",
            flower_grade=stock.flower_grade or "",
            supplier_code=stock.supplier_code or "",
            supplier_name=stock.supplier_name or "-",
            received_date=stock.received_date,
            quantity=quantity,
            reference_code=production.production_code,
            business_date=get_business_date_for_packing(),
            note=note,
        )

    # KIỂM TRA TRƯỚC, CHƯA TẠO THÙNG
    for item in production.items.all():
        qty = float(request.POST.get(f"packing_quantity_{item.id}") or 0)

        if qty <= 0:
            continue

        packed_qty = item.box_items.aggregate(total=Sum("stems"))["total"] or 0
        remaining_qty = max(float(item.required_quantity or 0) - float(packed_qty or 0), 0)

        if qty > remaining_qty:
            messages.error(request, f"{item.product_name} chỉ còn {remaining_qty} cành để đóng.")
            return redirect("production:packing_detail", production_id=production.id)

        finished_stock = None
        sale_stock = None

        standard_qty = float(item.standard_quantity or 0)
        sale_qty = float(item.sale_quantity or 0)

        already_packed = item.box_items.aggregate(total=Sum("stems"))["total"] or 0

        remaining_standard = max(standard_qty - already_packed, 0)
        remaining_sale = max(sale_qty - max(already_packed - standard_qty, 0), 0)

        deduct_finished_qty = min(qty, remaining_standard)
        deduct_sale_qty = qty - deduct_finished_qty

        if deduct_sale_qty > remaining_sale:
            messages.error(
                request,
                f"{item.product_name} số lượng sale chỉ còn {remaining_sale} cành để đóng."
            )
            return redirect("production:packing_detail", production_id=production.id)

        if deduct_finished_qty > 0:
            finished_stock = get_available_stock(
                warehouse_type="FINISHED",
                product_code=item.product_code,
                product_name=item.product_name,
            )

            if not finished_stock or float(finished_stock.quantity or 0) < deduct_finished_qty:
                available = finished_stock.quantity if finished_stock else 0
                messages.error(
                    request,
                    f"{item.product_name} kho thành phẩm chỉ còn {available} cành."
                )
                return redirect("production:packing_detail", production_id=production.id)

        if deduct_sale_qty > 0:
            sale_stock = get_available_stock(
                warehouse_type="SALE",
                product_code=item.product_code,
                product_name=item.product_name,
            )

            if not sale_stock or float(sale_stock.quantity or 0) < deduct_sale_qty:
                available = sale_stock.quantity if sale_stock else 0
                messages.error(
                    request,
                    f"{item.product_name} kho sale chỉ còn {available} cành."
                )
                return redirect("production:packing_detail", production_id=production.id)

        items_to_pack.append((
            item,
            qty,
            finished_stock,
            deduct_finished_qty,
            sale_stock,
            deduct_sale_qty,
        ))

    if not items_to_pack:
        messages.error(request, "Vui lòng nhập số lượng cho ít nhất một sản phẩm.")
        return redirect("production:packing_detail", production_id=production.id)

    # SAU KHI KIỂM TRA ĐỦ MỚI TẠO THÙNG
    box_number = production.boxes.count() + 1
    box_label = f"T{box_number:02d}"

    box = PackingBox.objects.create(
        production_order=production,
        box_code=box_label,
        box_number=box_number,
        box_type=box_type,
        capacity_index=box_type.capacity_index,
        status="draft",
    )

    for item, qty, finished_stock, deduct_finished_qty, sale_stock, deduct_sale_qty in items_to_pack:
        bunches = qty / stems_per_bunch

        packing_index_obj = get_packing_index_by_item(item)
        packing_index = packing_index_obj.packing_index if packing_index_obj else 0

        box_lot = None

        if finished_stock is not None and finished_stock.lot:
            box_lot = finished_stock.lot
        elif sale_stock is not None and sale_stock.lot:
            box_lot = sale_stock.lot
        else:
            box_lot = item.lot

        PackingBoxItem.objects.create(
            box=box,
            production_item=item,
            lot=box_lot,
            product_code=item.product_code,
            product_name=item.product_name,
            bunches=bunches,
            stems=qty,
            stems_per_bunch=stems_per_bunch,
            packing_index=packing_index,
            nw=0,
            gw=0,
        )

        if finished_stock is not None and deduct_finished_qty > 0:
            finished_stock.quantity = float(finished_stock.quantity or 0) - deduct_finished_qty
            finished_stock.save()

            create_packing_out_transaction(
                stock=finished_stock,
                quantity=deduct_finished_qty,
                box=box,
                production=production,
                note=(
                    f"Xuất kho thành phẩm do đóng gói sản xuất. "
                    f"Thùng {box.box_code} - Lệnh {production.production_code} - "
                    f"Đơn {production.order_code}"
                ),
            )

            trace_lot = finished_stock.lot or item.lot

            if trace_lot:
                add_trace_log(
                    lot=trace_lot,
                    action="packing",
                    quantity=deduct_finished_qty,
                    from_area="Kho thành phẩm",
                    to_area=f"Thùng {box.box_code}",
                    employee_name="",
                    related_code=box.box_code,
                    note=f"Đóng hàng tiêu chuẩn vào thùng {box.box_code} - Lệnh {production.production_code}",
                )

        if sale_stock is not None and deduct_sale_qty > 0:
            sale_stock.quantity = float(sale_stock.quantity or 0) - deduct_sale_qty
            sale_stock.save()

            create_packing_out_transaction(
                stock=sale_stock,
                quantity=deduct_sale_qty,
                box=box,
                production=production,
                note=(
                    f"Xuất kho sale do đóng gói sản xuất. "
                    f"Thùng {box.box_code} - Lệnh {production.production_code} - "
                    f"Đơn {production.order_code}"
                ),
            )

            trace_lot = sale_stock.lot or item.lot

            if trace_lot:
                add_trace_log(
                    lot=trace_lot,
                    action="packing",
                    quantity=deduct_sale_qty,
                    from_area="Kho Sale",
                    to_area=f"Thùng {box.box_code}",
                    employee_name="",
                    related_code=box.box_code,
                    note=f"Đóng hàng sale vào thùng {box.box_code} - Lệnh {production.production_code}",
                )

    recalculate_box(box)

    remaining_items = []

    for item in production.items.all():
        packed_qty = item.box_items.aggregate(total=Sum("stems"))["total"] or 0
        remaining_qty = max(float(item.required_quantity or 0) - float(packed_qty or 0), 0)

        if remaining_qty > 0:
            remaining_items.append(item.id)

    production.status = "packing" if remaining_items else "completed"

    if not remaining_items:
        production.completed_at = timezone.now()

    production.save()

    messages.success(request, f"Đã tạo thùng {box.box_code}.")
    return redirect("production:packing_detail", production_id=production.id)


def packing_list_view(request, production_id):
    production = get_object_or_404(
        ProductionOrder.objects.prefetch_related(
            "boxes",
            "boxes__items",
            "boxes__box_type",
            "order",
        ),
        id=production_id
    )

    boxes = list(
        production.boxes.all().order_by("box_number")
    )

    total_boxes = len(boxes)

    total_bunches = sum(
        box.total_bunches or 0
        for box in boxes
    )

    total_stems = sum(
        box.total_stems or 0
        for box in boxes
    )

    total_nw = sum(
        box.nw or 0
        for box in boxes
    )

    total_gw = sum(
        box.gw or 0
        for box in boxes
    )

    pkl_pages = []

    for box in boxes:
        qr_url = request.build_absolute_uri(
            f"/production/box/{box.id}/"
        )

        qr = qrcode.make(qr_url)

        buffer = BytesIO()
        qr.save(buffer, format="PNG")

        qr_base64 = base64.b64encode(
            buffer.getvalue()
        ).decode()

        items = list(
            box.items.all().order_by("id")
        )

        half = (len(items) + 1) // 2

        left_items = items[:half]
        right_items = items[half:]

        pkl_pages.append({
            "box": box,
            "items": items,
            "left_items": left_items,
            "right_items": right_items,
            "qr_code": qr_base64,
        })

    box_pairs = []

    for i in range(0, len(boxes), 2):

        left_box = boxes[i]

        right_box = (
            boxes[i + 1]
            if i + 1 < len(boxes)
            else None
        )

        box_pairs.append({
            "left": left_box,
            "right": right_box,
        })

    return render(
        request,
        "production/packing_list.html",
        {
            "production": production,
            "boxes": boxes,

            # PKL mới
            "pkl_pages": pkl_pages,

            # dữ liệu cũ giữ nguyên
            "box_pairs": box_pairs,
            "total_boxes": total_boxes,
            "total_bunches": total_bunches,
            "total_stems": total_stems,
            "total_nw": total_nw,
            "total_gw": total_gw,
        }
    )


def export_packing_list_excel(request, production_id):
    production = get_object_or_404(
        ProductionOrder.objects.prefetch_related("boxes__items"),
        id=production_id
    )

    template_path = os.path.join(
        settings.BASE_DIR,
        "data",
        "templates",
        "packing_list_template.xlsx"
    )

    wb = load_workbook(template_path)
    ws_pkl = wb["SHOW PKL"]

    ws_pkl["B3"] = production.customer_name
    ws_pkl["B4"] = production.order_code
    ws_pkl["B5"] = production.production_code

    row = 10

    for box in production.boxes.all().order_by("box_number"):
        for item in box.items.all():
            ws_pkl.cell(row=row, column=1).value = box.box_number
            ws_pkl.cell(row=row, column=2).value = box.box_code
            ws_pkl.cell(row=row, column=3).value = item.product_name
            ws_pkl.cell(row=row, column=4).value = item.product_code
            ws_pkl.cell(row=row, column=5).value = item.bunches
            ws_pkl.cell(row=row, column=6).value = item.stems_per_bunch
            ws_pkl.cell(row=row, column=7).value = item.stems
            ws_pkl.cell(row=row, column=8).value = item.nw
            ws_pkl.cell(row=row, column=9).value = box.gw
            row += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    filename = f"SHOW_PKL_{production.order_code}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

def packing_box_detail(request, box_id):

    box = get_object_or_404(
        PackingBox.objects.prefetch_related(
            "items",
            "production_order",
            "production_order__order",
        ),
        id=box_id
    )

    qr_data = f"{box.box_code} | {box.production_order.order_code}"

    qr_img = qrcode.make(qr_data)

    buffer = BytesIO()
    qr_img.save(buffer, format="PNG")

    qr_code = base64.b64encode(
        buffer.getvalue()
    ).decode("utf-8")

    items = list(box.items.all())

    left_items = []
    right_items = []

    for index, item in enumerate(items):

        if index % 2 == 0:
            left_items.append(item)

        else:
            right_items.append(item)

    return render(request, "production/packing_box_detail.html", {
        "box": box,
        "items": items,
        "left_items": left_items,
        "right_items": right_items,
        "qr_code": qr_code,
    })

def edit_packing_box(request, box_id):
    box = get_object_or_404(PackingBox, id=box_id)

    production_id = box.production_order.id
    box_code = box.box_code

    box.delete()

    messages.info(
        request,
        f"Đã mở lại {box_code}. Vui lòng tạo lại thùng với số lượng đúng."
    )

    return redirect("production:packing_detail", production_id=production_id)


def delete_packing_box(request, box_id):
    box = get_object_or_404(PackingBox, id=box_id)

    production_id = box.production_order.id

    box.delete()

    messages.success(request, "Đã xóa thùng.")

    return redirect("production:packing_detail", production_id=production_id)


def reset_packing(request, production_id):
    production = get_object_or_404(ProductionOrder, id=production_id)

    production.boxes.all().delete()
    production.status = "packing"
    production.save()

    messages.info(request, f"Đã mở lại đóng gói cho lệnh {production.production_code}. Vui lòng đóng gói lại.")

    return redirect("production:packing_detail", production_id=production.id)


def export_packing_list_pdf(request, production_id):
    response_html = packing_list_view(request, production_id)
    html = response_html.content.decode("utf-8")

    pdf_file = BytesIO()

    pisa.CreatePDF(
        html,
        dest=pdf_file
    )

    response = HttpResponse(
        pdf_file.getvalue(),
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        f'inline; filename="PKL-{production_id}.pdf"'
    )

    return response

def packing_label_view(request, box_id):
    box = get_object_or_404(
        PackingBox.objects.select_related(
            "production_order",
            "production_order__order",
        ).prefetch_related("items"),
        id=box_id
    )

    qr_url = request.build_absolute_uri(
        f"/production/box/{box.id}/"
    )

    qr = qrcode.make(qr_url)
    buffer = BytesIO()
    qr.save(buffer, format="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()

    items = list(box.items.all())
    half = (len(items) + 1) // 2

    return render(request, "production/packing_label.html", {
        "box": box,
        "production": box.production_order,
        "left_items": items[:half],
        "right_items": items[half:],
        "qr_code": qr_base64,
        "label_copies": [1, 2],
    })

def pending_orders_voice_api(request):
    received_order_ids = ProductionOrder.objects.values_list("order_id", flat=True)

    pending_orders = (
        Order.objects
        .prefetch_related("items")
        .exclude(id__in=received_order_ids)
        .order_by("-order_time")[:10]
    )

    data = []

    for order in pending_orders:
        product_names = [item.product_name for item in order.items.all()]

        voice_text = (
            f"Thông báo đơn hàng mới. "
            f"Khách hàng {order.customer_name}. "
            f"Mã đơn {order.order_code}. "
            f"Đơn hàng có {order.items.count()} sản phẩm. "
            f"Danh sách gồm: {', '.join(product_names)}. "
            f"Bộ phận sản xuất vui lòng chuẩn bị hàng."
        )

        data.append({
            "order_code": order.order_code,
            "customer_name": order.customer_name,
            "items_count": order.items.count(),
            "voice_text": voice_text,
        })

    return JsonResponse({
        "orders": data
    })

def suggest_box_api(request):
    product_code = request.GET.get("product_code", "")
    product_name = request.GET.get("product_name", "")
    stems = float(request.GET.get("stems") or 0)

    packing_index_obj = None

    if product_code:
        packing_index_obj = PackingIndex.objects.filter(
            flower__code__iexact=product_code,
            is_active=True
        ).first()

    if not packing_index_obj and product_name:
        packing_index_obj = PackingIndex.objects.filter(
            flower__name__icontains=product_name,
            is_active=True
        ).first()

    packing_index = packing_index_obj.packing_index if packing_index_obj else 0
    base_stems = packing_index_obj.base_stems if packing_index_obj else 50

    total_index = 0
    if base_stems > 0:
        total_index = (stems / base_stems) * packing_index

    suggested_box = (
        BoxCapacity.objects
        .filter(is_active=True, capacity_index__gte=total_index)
        .order_by("capacity_index")
        .first()
    )

    if not suggested_box:
        suggested_box = (
            BoxCapacity.objects
            .filter(is_active=True)
            .order_by("-capacity_index")
            .first()
        )

    return JsonResponse({
        "product_code": product_code,
        "product_name": product_name,
        "stems": stems,
        "packing_index": packing_index,
        "base_stems": base_stems,
        "total_index": total_index,
        "box_id": suggested_box.id if suggested_box else None,
        "box_code": suggested_box.code if suggested_box else "",
        "box_name": suggested_box.name if suggested_box else "",
        "capacity_index": suggested_box.capacity_index if suggested_box else 0,
    })